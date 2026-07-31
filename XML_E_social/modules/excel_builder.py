from __future__ import annotations

import csv
import sqlite3
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable, Iterator, Sequence

import pandas as pd
from openpyxl import Workbook, load_workbook


ProgressCallback = Callable[[float, str], None]
MAX_DADOS_ABA = 1_048_575
CHUNK_EXPORT = 25_000
CARACTERES_INVALIDOS_ABA = set(r"[]:*?/\\")


@dataclass(frozen=True)
class FontePlanilha:
    nome: str
    dataframe: pd.DataFrame | None = None
    query: str | None = None
    params: tuple = ()
    total_esperado: int | None = None


@dataclass
class ControleAba:
    conjunto: str
    aba: str
    parte: int
    linhas_esperadas: int
    linhas_exportadas: int
    status: str


@dataclass
class ResultadoWorkbook:
    caminho: str
    abas: list[str]
    controles: list[ControleAba] = field(default_factory=list)
    manifesto: str | None = None


def _progresso(
    callback: ProgressCallback | None, valor: float, mensagem: str
) -> None:
    if callback:
        callback(max(0.0, min(1.0, valor)), mensagem)


def _nome_aba(nome: str, parte: int, total_partes: int, usados: set[str]) -> str:
    base = "".join("_" if c in CARACTERES_INVALIDOS_ABA else c for c in nome).strip()
    base = base or "Planilha"
    sufixo = f"_{parte}" if total_partes > 1 else ""
    candidato = f"{base[:31-len(sufixo)]}{sufixo}"
    contador = 2
    while candidato.casefold() in usados:
        extra = f"_{contador}"
        candidato = f"{base[:31-len(sufixo)-len(extra)]}{sufixo}{extra}"
        contador += 1
    usados.add(candidato.casefold())
    return candidato


def _valor_excel(valor: object) -> object:
    if valor is None:
        return None
    if isinstance(valor, float) and pd.isna(valor):
        return None
    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime().replace(tzinfo=None)
    if isinstance(valor, (datetime, date, str, int, float, bool)):
        return valor
    return str(valor)


def _iter_dataframe(
    dataframe: pd.DataFrame,
) -> tuple[list[str], int, Iterator[Sequence[object]]]:
    colunas = [str(c) for c in dataframe.columns]
    return colunas, len(dataframe), dataframe.itertuples(index=False, name=None)


def _iter_query(
    conn: sqlite3.Connection, query: str, params: tuple, chunk: int
) -> tuple[list[str], Iterator[Sequence[object]]]:
    cursor = conn.execute(query, params)
    colunas = [str(d[0]) for d in cursor.description]

    def linhas() -> Iterator[Sequence[object]]:
        try:
            while True:
                bloco = cursor.fetchmany(chunk)
                if not bloco:
                    return
                yield from bloco
        finally:
            cursor.close()

    return colunas, linhas()


def _total_query(conn: sqlite3.Connection, query: str, params: tuple) -> int:
    row = conn.execute(f"SELECT COUNT(*) FROM ({query}) AS fonte_v95", params).fetchone()
    return int(row[0] if row else 0)


def _escrever_fonte(
    workbook: Workbook,
    fonte: FontePlanilha,
    conn: sqlite3.Connection | None,
    usados: set[str],
    callback: ProgressCallback | None,
    inicio: float,
    fim: float,
    max_dados_aba: int,
    chunk: int,
) -> list[ControleAba]:
    if fonte.query is not None:
        if conn is None:
            raise ValueError(f"A fonte '{fonte.nome}' exige uma conexao SQLite.")
        total = (
            fonte.total_esperado
            if fonte.total_esperado is not None
            else _total_query(conn, fonte.query, fonte.params)
        )
        colunas, linhas = _iter_query(conn, fonte.query, fonte.params, chunk)
    else:
        dataframe = fonte.dataframe if fonte.dataframe is not None else pd.DataFrame()
        colunas, total, linhas = _iter_dataframe(dataframe)

    total_partes = max(1, (total + max_dados_aba - 1) // max_dados_aba)
    controles: list[ControleAba] = []
    planilha = None
    exportadas_parte = 0
    exportadas_total = 0
    parte = 0

    def nova_planilha() -> None:
        nonlocal planilha, exportadas_parte, parte
        if planilha is not None:
            controles.append(
                ControleAba(
                    conjunto=fonte.nome,
                    aba=planilha.title,
                    parte=parte,
                    linhas_esperadas=min(
                        max_dados_aba, max(total - (parte - 1) * max_dados_aba, 0)
                    ),
                    linhas_exportadas=exportadas_parte,
                    status="OK",
                )
            )
        parte += 1
        exportadas_parte = 0
        planilha = workbook.create_sheet(
            _nome_aba(fonte.nome, parte, total_partes, usados)
        )
        if colunas:
            planilha.append(colunas)

    nova_planilha()
    for linha in linhas:
        if exportadas_parte >= max_dados_aba:
            nova_planilha()
        planilha.append([_valor_excel(valor) for valor in linha])
        exportadas_parte += 1
        exportadas_total += 1
        if exportadas_total % chunk == 0:
            _progresso(
                callback,
                inicio + (fim - inicio) * exportadas_total / max(total, 1),
                (
                    f"Exportando {fonte.nome}: "
                    f"{exportadas_total:,} de {total:,}"
                ).replace(",", "."),
            )
    controles.append(
        ControleAba(
            conjunto=fonte.nome,
            aba=planilha.title,
            parte=parte,
            linhas_esperadas=min(
                max_dados_aba, max(total - (parte - 1) * max_dados_aba, 0)
            ),
            linhas_exportadas=exportadas_parte,
            status="OK" if exportadas_total == total else "REVISAR",
        )
    )
    if exportadas_total != total:
        raise RuntimeError(
            f"Integridade de '{fonte.nome}': esperado={total}, exportado={exportadas_total}."
        )
    return controles


def _escrever_controle(
    workbook: Workbook, controles: Iterable[ControleAba], usados: set[str]
) -> None:
    ws = workbook.create_sheet(
        _nome_aba("controle_integridade", 1, 1, usados)
    )
    ws.append(
        [
            "conjunto",
            "aba",
            "parte",
            "linhas_esperadas",
            "linhas_exportadas",
            "status",
        ]
    )
    for controle in controles:
        ws.append(
            [
                controle.conjunto,
                controle.aba,
                controle.parte,
                controle.linhas_esperadas,
                controle.linhas_exportadas,
                controle.status,
            ]
        )


def validar_workbook(
    caminho: str | Path, controles: Sequence[ControleAba]
) -> list[str]:
    arquivo = Path(caminho)
    if not arquivo.is_file() or arquivo.stat().st_size == 0:
        raise RuntimeError(f"Workbook ausente ou vazio: {arquivo}")
    with zipfile.ZipFile(arquivo) as pacote:
        corrompido = pacote.testzip()
        if corrompido:
            raise RuntimeError(f"Entrada corrompida no XLSX: {corrompido}")

    workbook = load_workbook(arquivo, read_only=True, data_only=False)
    try:
        if len(workbook.sheetnames) != len(set(n.casefold() for n in workbook.sheetnames)):
            raise RuntimeError("O workbook possui nomes de abas duplicados.")
        if "controle_integridade" not in workbook.sheetnames:
            raise RuntimeError("A aba controle_integridade nao foi encontrada.")
        for controle in controles:
            if controle.aba not in workbook.sheetnames:
                raise RuntimeError(f"Aba ausente: {controle.aba}")
            planilha = workbook[controle.aba]
            total_linhas_fisicas = sum(1 for _ in planilha.iter_rows(values_only=True))
            linhas = max(total_linhas_fisicas - 1, 0)
            if linhas != controle.linhas_exportadas:
                raise RuntimeError(
                    f"Integridade da aba {controle.aba}: "
                    f"controle={controle.linhas_exportadas}, workbook={linhas}."
                )
            if linhas > MAX_DADOS_ABA:
                raise RuntimeError(f"Aba {controle.aba} excedeu o limite do Excel.")
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def gerar_workbook(
    caminho_saida: str | Path,
    fontes: Sequence[FontePlanilha],
    *,
    conexao: sqlite3.Connection | None = None,
    progress_callback: ProgressCallback | None = None,
    gerar_manifesto: bool = False,
    max_dados_aba: int = MAX_DADOS_ABA,
    chunk: int = CHUNK_EXPORT,
) -> ResultadoWorkbook:
    destino = Path(caminho_saida).expanduser().resolve()
    destino.parent.mkdir(parents=True, exist_ok=True)
    parcial = destino.with_name(f"{destino.stem}.parcial{destino.suffix}")
    workbook = Workbook(write_only=True)
    usados: set[str] = set()
    controles: list[ControleAba] = []
    total_fontes = max(len(fontes), 1)

    _progresso(progress_callback, 0.01, "Preparando workbook consolidado...")
    for indice, fonte in enumerate(fontes):
        inicio = 0.02 + 0.90 * indice / total_fontes
        fim = 0.02 + 0.90 * (indice + 1) / total_fontes
        controles.extend(
            _escrever_fonte(
                workbook,
                fonte,
                conexao,
                usados,
                progress_callback,
                inicio,
                fim,
                max_dados_aba,
                chunk,
            )
        )
    _escrever_controle(workbook, controles, usados)
    _progresso(progress_callback, 0.94, "Salvando workbook consolidado...")
    workbook.save(parcial)
    _progresso(progress_callback, 0.97, "Validando workbook e integridade...")
    abas = validar_workbook(parcial, controles)
    parcial.replace(destino)

    manifesto_path: Path | None = None
    if gerar_manifesto:
        manifesto_path = destino.with_name(f"{destino.stem}_manifesto.csv")
        with manifesto_path.open("w", newline="", encoding="utf-8-sig") as arquivo:
            writer = csv.DictWriter(
                arquivo,
                fieldnames=[
                    "conjunto",
                    "aba",
                    "parte",
                    "linhas_esperadas",
                    "linhas_exportadas",
                    "status",
                ],
                delimiter=";",
            )
            writer.writeheader()
            writer.writerows(vars(controle) for controle in controles)
    _progresso(progress_callback, 1.0, "Relatorio consolidado validado.")
    return ResultadoWorkbook(
        caminho=str(destino),
        abas=abas,
        controles=controles,
        manifesto=str(manifesto_path) if manifesto_path else None,
    )
