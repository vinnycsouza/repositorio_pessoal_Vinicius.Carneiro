from __future__ import annotations

import math
import os
import pickle
import shutil
import tempfile
import sqlite3
import time
import zlib
from dataclasses import asdict
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd
from openpyxl import Workbook

from modules.auditoria import (
    CODIGOS_CP_EXPORTACAO_PADRAO,
    _observacao_rubrica,
    _prioridade_rubrica,
    classificar_carater,
    classificar_status_cp,
    classificar_tipo_verba,
    entra_base_cp,
)
from modules.excel_builder import FontePlanilha, gerar_workbook
from modules.busca_rubricas import combinar_filtros_sql, filtro_sql_chaves_rubricas
from modules.progresso import emitir_progresso
from modules.sqlite_writer import BatchPolicy, SQLiteWriter

ProgressCallback = Callable[[float, str], None]
MAX_DADOS_ABA = 1_048_575
CHUNK_EXPORT = 25_000


_SQL_MOVIMENTOS_CP_EFETIVOS = """
SELECT r.*
FROM dados_remuneracoes AS r
CROSS JOIN eventos AS e ON e.arquivo = r.arquivo
WHERE e.tipo='S-1200'
  AND e.ativo=1
  AND e.duplicado_logico_de IS NULL
  AND NOT EXISTS (
    SELECT 1
    FROM dados_exclusoes AS x
    WHERE COALESCE(x.nrRecEvt,'')<>''
      AND x.nrRecEvt = CASE
        WHEN COALESCE(e.recibo_evento,'')<>'' THEN e.recibo_evento
        WHEN COALESCE(e.ind_retif,'')<>'2'
         AND COALESCE(r.nr_recibo_evento,'')<>COALESCE(e.recibo_referencia,'')
          THEN COALESCE(r.nr_recibo_evento,'')
        ELSE ''
      END
  )
"""


SCHEMAS_PADRAO = {
    "dados_rubricas": {"cod_rubr":"", "ide_tab_rubr":"", "dsc_rubr":"", "nat_rubr":"", "cod_inc_cp":"", "cod_inc_fgts":"", "cod_inc_irrf":"", "tp_rubr":"", "origem_bloco":"", "ini_valid":"", "fim_valid":"", "arquivo_origem":"", "fonte_dados":""},
    "dados_exclusoes": {"nrRecEvt":"", "arquivo_origem":""},
    "dados_remuneracoes": {"arquivo":"", "cpf":"", "matricula":"", "per_apur":"", "cod_categ":"", "tp_insc_estab":"", "nr_insc_estab":"", "cod_lotacao":"", "cod_rubr":"", "ide_tab_rubr":"", "vr_rubr":0.0, "nat_rubr":"", "cod_inc_cp":"", "dsc_rubr":"", "tp_rubr":"", "ini_valid":"", "fim_valid":"", "origem_bloco_s1010":"", "fonte_s1010":"", "arquivo_s1010":"", "criterio_cruzamento_s1010":"", "origem_validacao":"", "nivel_confianca":"", "status_auditoria":"", "observacao_validacao":"", "nr_recibo_evento":"", "status_cp":"", "considerado_cp":"", "tipo_verba":"", "carater_verba":""},
    "dados_remuneracoes_s2299": {"arquivo":"", "cpf":"", "matricula":"", "per_apur":"", "data_desligamento":"", "ide_dm_dev":"", "cod_categ":"", "tp_insc_estab":"", "nr_insc_estab":"", "cod_lotacao":"", "cod_rubr":"", "ide_tab_rubr":"", "vr_rubr":0.0, "nat_rubr":"", "cod_inc_cp":"", "dsc_rubr":"", "tp_rubr":"", "ini_valid":"", "fim_valid":"", "criterio_cruzamento_s1010":"", "origem_validacao":"", "nivel_confianca":"", "status_auditoria":"", "observacao_validacao":"", "nr_recibo_evento":"", "origem_movimento":"S-2299", "status_cp":"", "considerado_cp":"", "tipo_verba":"", "carater_verba":""},
    "dados_bases_trabalhador": {"arquivo":"", "cpf":"", "matricula":"", "per_apur":"", "per_ref":"", "cod_categ":"", "tp_insc_estab":"", "nr_insc_estab":"", "cod_lotacao":"", "ind13":"", "tp_valor":"", "valor":0.0, "origem_valor":"", "nr_recibo_base":""},
    "dados_bases_contribuicao": {"arquivo":"", "per_apur":"", "tp_insc_estab":"", "nr_insc_estab":"", "cod_lotacao":"", "cod_categ":"", "ind_incid":"", "fpas":"", "cod_tercs":"", "aliq_rat_ajust":0.0, "vr_bc_cp":0.0, "vr_bc_cp_00":0.0, "vr_bc_cp_15":0.0, "vr_bc_cp_20":0.0, "vr_bc_cp_25":0.0, "nr_recibo_base":""},
    "dados_empresa": {"nome_empresa":"", "cnpj_empregador":""},
    "dados_contexto_empregador": {"arquivo":"", "tp_insc":"", "nr_insc":"", "operacao":"", "ini_valid":"", "fim_valid":"", "classificacao_tributaria":"", "indicador_desoneracao":"", "nome_empresa":""},
    "dados_contexto_lotacao": {"arquivo":"", "tp_insc":"", "nr_insc":"", "cod_lotacao":"", "operacao":"", "ini_valid":"", "fim_valid":"", "tp_lotacao":"", "fpas":"", "cod_tercs":""},
    "dados_contexto_trabalhador": {"arquivo":"", "cpf":"", "matricula":"", "data_admissao":"", "cod_categ":"", "tp_reg_trab":"", "tp_reg_prev":"", "cod_cargo":"", "cod_funcao":"", "cbo_cargo":""},
}


def _progresso(cb: ProgressCallback | None, valor: float, msg: str) -> None:
    if cb:
        cb(max(0.0, min(1.0, valor)), msg)


def _q(nome: str) -> str:
    return '"' + nome.replace('"', '""') + '"'


def _sql_type(valor: object) -> str:
    return "REAL" if isinstance(valor, (int, float)) and not isinstance(valor, bool) else "TEXT"


def _obj_dict(obj: object) -> dict:
    if hasattr(obj, "__dataclass_fields__"):
        return asdict(obj)
    if isinstance(obj, dict):
        return dict(obj)
    return {"valor": str(obj)}


def _iter_payloads(conn: sqlite3.Connection, categoria: str, apos_id: int = 0):
    cur = conn.execute(
        "SELECT id, evento_id, payload FROM objetos WHERE categoria=? AND id>? ORDER BY id",
        (categoria, apos_id),
    )
    for obj_id, evento_id, payload in cur:
        bloco = pickle.loads(zlib.decompress(payload))
        itens = bloco if isinstance(bloco, list) else [bloco]
        yield int(obj_id), evento_id, itens


def _criar_tabela_por_amostra(conn: sqlite3.Connection, tabela: str, amostra: dict) -> list[str]:
    colunas = list(amostra.keys())
    defs = ", ".join(f"{_q(c)} {_sql_type(amostra[c])}" for c in colunas)
    conn.execute(f"CREATE TABLE IF NOT EXISTS {_q(tabela)} ({defs})")
    return colunas


def _colunas_tabela(conn: sqlite3.Connection, tabela: str) -> list[str]:
    return [r[1] for r in conn.execute(f"PRAGMA table_info({_q(tabela)})")]


def materializar_tabelas_analiticas(
    conn: sqlite3.Connection,
    progress_callback: ProgressCallback | None = None,
) -> None:
    """Converte os BLOBs por evento em tabelas relacionais sem carregar tudo na RAM.

    A operação é retomável: o último id de payload processado é salvo em meta.
    """
    categorias = [
        ("rubricas", "dados_rubricas"),
        ("exclusoes", "dados_exclusoes"),
        ("remuneracoes", "dados_remuneracoes"),
        ("remuneracoes_s2299", "dados_remuneracoes_s2299"),
        ("bases_trabalhador", "dados_bases_trabalhador"),
        ("bases_contribuicao", "dados_bases_contribuicao"),
        ("empresa", "dados_empresa"),
        ("contexto_empregador", "dados_contexto_empregador"),
        ("contexto_lotacao", "dados_contexto_lotacao"),
        ("contexto_trabalhador", "dados_contexto_trabalhador"),
    ]
    escritor = SQLiteWriter(conn, BatchPolicy.atual("analitico"))
    for tabela, amostra in SCHEMAS_PADRAO.items():
        if not _colunas_tabela(conn, tabela):
            _criar_tabela_por_amostra(conn, tabela, amostra)
    conn.commit()

    totais = {
        cat: int(conn.execute("SELECT COUNT(*) FROM objetos WHERE categoria=?", (cat,)).fetchone()[0])
        for cat, _ in categorias
    }
    total_payloads = max(sum(totais.values()), 1)
    feitos_global = 0
    emitir_progresso(
        progress_callback, "materializacao", 0.0,
        "Preparando tabelas analíticas no SQLite...",
        f"{total_payloads:,} blocos persistidos serão verificados.".replace(",", "."),
    )

    for categoria, tabela in categorias:
        chave = f"materializado_{categoria}_ate"
        row = conn.execute("SELECT valor FROM meta WHERE chave=?", (chave,)).fetchone()
        ultimo = int(row[0]) if row and str(row[0]).isdigit() else 0
        feitos_global += int(conn.execute(
            "SELECT COUNT(*) FROM objetos WHERE categoria=? AND id<=?", (categoria, ultimo)
        ).fetchone()[0])
        colunas = _colunas_tabela(conn, tabela)
        buffer: list[tuple] = []
        bytes_buffer = 0
        payloads_desde_commit = 0

        for obj_id, evento_id, itens in _iter_payloads(conn, categoria, ultimo):
            for obj in itens:
                d = _obj_dict(obj)
                d.setdefault("evento_id_origem", evento_id)
                if categoria in {"remuneracoes", "remuneracoes_s2299"}:
                    d["status_cp"] = classificar_status_cp(d.get("cod_inc_cp", ""))
                    d["considerado_cp"] = "Sim" if entra_base_cp(d.get("cod_inc_cp", "")) else "Não"
                    d["tipo_verba"] = classificar_tipo_verba(d.get("dsc_rubr", ""), d.get("nat_rubr", ""), d.get("tp_rubr", ""))
                    d["carater_verba"] = classificar_carater(d.get("dsc_rubr", ""), d.get("nat_rubr", ""), d.get("tp_rubr", ""))
                if not colunas:
                    colunas = _criar_tabela_por_amostra(conn, tabela, d)
                # Compatibilidade com campos novos sem quebrar workspaces antigos.
                faltantes = [c for c in d if c not in colunas]
                for c in faltantes:
                    conn.execute(f"ALTER TABLE {_q(tabela)} ADD COLUMN {_q(c)} {_sql_type(d[c])}")
                    colunas.append(c)
                linha = tuple(d.get(c, None) for c in colunas)
                buffer.append(linha)
                bytes_buffer += sum(
                    len(str(valor).encode("utf-8", "replace"))
                    for valor in linha if valor is not None
                )

            payloads_desde_commit += 1
            feitos_global += 1
            if (
                len(buffer) >= escritor.policy.max_itens
                or bytes_buffer >= escritor.policy.max_bytes
                or payloads_desde_commit >= 250
            ):
                marks = ",".join("?" for _ in colunas)
                escritor.executemany(
                    f"INSERT INTO {_q(tabela)} ({','.join(_q(c) for c in colunas)}) VALUES ({marks})",
                    buffer,
                )
                buffer.clear()
                bytes_buffer = 0
                conn.execute(
                    "INSERT INTO meta(chave,valor) VALUES(?,?) ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor",
                    (chave, str(obj_id)),
                )
                conn.execute(
                    "INSERT INTO meta(chave,valor) VALUES('atualizado_em',?) ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor",
                    (str(time.time()),),
                )
                escritor.commit()
                payloads_desde_commit = 0
                emitir_progresso(
                    progress_callback, "materializacao", feitos_global / total_payloads,
                    f"Materialização segura: {categoria} — {feitos_global:,} blocos persistidos".replace(",", "."),
                    f"{feitos_global:,} de {total_payloads:,} blocos persistidos".replace(",", "."),
                )
        if buffer:
            marks = ",".join("?" for _ in colunas)
            escritor.executemany(
                f"INSERT INTO {_q(tabela)} ({','.join(_q(c) for c in colunas)}) VALUES ({marks})",
                buffer,
            )
            conn.execute(
                "INSERT INTO meta(chave,valor) VALUES(?,?) ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor",
                (chave, str(obj_id)),
            )
            escritor.commit()

    emitir_progresso(
        progress_callback, "materializacao", 0.98,
        "Criando índices para acelerar as consultas...",
    )
    _criar_indices(conn)
    escritor.registrar_telemetria()
    conn.commit()
    emitir_progresso(
        progress_callback, "materializacao", 1.0,
        "Materialização analítica concluída.",
    )
    _criar_tabelas_relatorio(conn, progress_callback)


def reiniciar_materializacao_analitica(conn: sqlite3.Connection) -> None:
    """Descarta somente projeções derivadas para reconstrução a partir de objetos.

    Os eventos/XML e os objetos persistidos permanecem intactos. Esta operação é
    necessária quando um S-1010 novo pode mudar a classificação de S-1200 antigos.
    """
    tabelas = [
        "dados_rubricas", "dados_exclusoes", "dados_remuneracoes",
        "dados_remuneracoes_s2299", "dados_contexto_empregador",
        "dados_contexto_lotacao", "dados_contexto_trabalhador",
        "dados_bases_trabalhador", "dados_bases_contribuicao", "dados_empresa",
        "rel_movimentos_cp", "rel_movimentos_s2299", "rel_rubricas_cp_base", "rel_sem_s1010",
        "rel_s5001_resumo", "rel_base_trabalhador", "rel_controle_integridade",
        "rel_reconciliacao_s5011",
    ]
    for tabela in tabelas:
        conn.execute(f"DROP TABLE IF EXISTS {_q(tabela)}")
    conn.execute("DELETE FROM meta WHERE chave LIKE 'materializado_%_ate'")
    conn.commit()


def _criar_indices(conn: sqlite3.Connection) -> None:
    comandos = [
        "CREATE INDEX IF NOT EXISTS idx_remun_arquivo ON dados_remuneracoes(arquivo)",
        "CREATE INDEX IF NOT EXISTS idx_remun_per ON dados_remuneracoes(per_apur)",
        "CREATE INDEX IF NOT EXISTS idx_remun_rubr ON dados_remuneracoes(cod_rubr, ide_tab_rubr)",
        "CREATE INDEX IF NOT EXISTS idx_remun_cp ON dados_remuneracoes(cod_inc_cp)",
        "CREATE INDEX IF NOT EXISTS idx_remun_trab ON dados_remuneracoes(per_apur, cpf, matricula)",
        "CREATE INDEX IF NOT EXISTS idx_remun_recibo ON dados_remuneracoes(nr_recibo_evento)",
        "CREATE INDEX IF NOT EXISTS idx_base_trab_chave ON dados_bases_trabalhador(per_apur, cpf, matricula, cod_categ, cod_lotacao)",
        "CREATE INDEX IF NOT EXISTS idx_base_trab_recibo ON dados_bases_trabalhador(nr_recibo_base)",
        "CREATE INDEX IF NOT EXISTS idx_excl_recibo ON dados_exclusoes(nrRecEvt)",
        "CREATE INDEX IF NOT EXISTS idx_s2299_arquivo ON dados_remuneracoes_s2299(arquivo)",
        "CREATE INDEX IF NOT EXISTS idx_s2299_per ON dados_remuneracoes_s2299(per_apur)",
        "CREATE INDEX IF NOT EXISTS idx_s2299_rubr ON dados_remuneracoes_s2299(cod_rubr, ide_tab_rubr)",
        "CREATE INDEX IF NOT EXISTS idx_contexto_lotacao_chave ON dados_contexto_lotacao(cod_lotacao,ini_valid)",
        "CREATE INDEX IF NOT EXISTS idx_contexto_trabalhador_chave ON dados_contexto_trabalhador(cpf,matricula)",
    ]
    for sql in comandos:
        try:
            conn.execute(sql)
        except sqlite3.OperationalError:
            pass
    conn.commit()


def _formatar_duracao_consulta(segundos: float) -> str:
    total = max(0, int(segundos))
    horas, resto = divmod(total, 3600)
    minutos, segundos = divmod(resto, 60)
    if horas:
        return f"{horas}h {minutos:02d}min"
    if minutos:
        return f"{minutos}min {segundos:02d}s"
    return f"{segundos}s"


def _executar_consulta_com_heartbeat(
    conn: sqlite3.Connection,
    sql: str,
    cb: ProgressCallback | None,
    progresso: float,
    mensagem: str,
) -> None:
    """Executa SQL longo mantendo o HUD vivo, sem consultar o banco no callback."""
    inicio = time.monotonic()
    ultimo_aviso = inicio

    def avisar_atividade() -> int:
        nonlocal ultimo_aviso
        agora = time.monotonic()
        if cb is not None and agora - ultimo_aviso >= 5.0:
            ultimo_aviso = agora
            try:
                emitir_progresso(
                    cb,
                    "consolidacao",
                    progresso,
                    mensagem,
                    "Consulta SQLite ativa há "
                    f"{_formatar_duracao_consulta(agora - inicio)}.",
                )
            except Exception:
                # O HUD é observabilidade: uma falha visual não deve cancelar
                # a transação SQLite que preserva a consistência do Workspace.
                pass
        return 0

    conn.set_progress_handler(avisar_atividade, 250_000)
    try:
        conn.execute(sql)
    finally:
        conn.set_progress_handler(None, 0)


def _recriar_tabela_derivada(
    conn: sqlite3.Connection,
    tabela: str,
    consulta: str,
    cb: ProgressCallback | None,
    progresso: float,
    mensagem: str,
) -> None:
    """Monta uma projeção ao lado da atual e a substitui somente ao concluir."""
    temporaria = f"{tabela}_nova"
    conn.execute(f"DROP TABLE IF EXISTS {_q(temporaria)}")
    conn.commit()
    _executar_consulta_com_heartbeat(
        conn,
        f"CREATE TABLE {_q(temporaria)} AS {consulta}",
        cb,
        progresso,
        mensagem,
    )
    conn.commit()
    with conn:
        conn.execute(f"DROP TABLE IF EXISTS {_q(tabela)}")
        conn.execute(
            f"ALTER TABLE {_q(temporaria)} RENAME TO {_q(tabela)}"
        )


def _criar_tabelas_relatorio(conn: sqlite3.Connection, cb: ProgressCallback | None) -> None:
    # Migrações de Workspace podem reconstruir derivados sem passar novamente
    # pela materialização. As tabelas novas da V10 precisam existir, ainda que
    # vazias, para manter compatibilidade com bancos V9.
    for tabela, amostra in SCHEMAS_PADRAO.items():
        if not _colunas_tabela(conn, tabela):
            _criar_tabela_por_amostra(conn, tabela, amostra)
    conn.commit()
    emitir_progresso(
        cb, "consolidacao", 0.0,
        "Consolidando rubricas e totais diretamente no SQLite...",
    )
    emitir_progresso(
        cb, "consolidacao", 0.05,
        "Selecionando movimentos S-1200 efetivos em uma única varredura...",
        "Os XMLs e as tabelas analíticas permanecem intactos.",
    )
    _recriar_tabela_derivada(
        conn,
        "rel_movimentos_cp",
        _SQL_MOVIMENTOS_CP_EFETIVOS,
        cb,
        0.05,
        "Selecionando movimentos S-1200 efetivos em uma única varredura...",
    )
    emitir_progresso(
        cb, "consolidacao", 0.35,
        "Movimentos efetivos selecionados; criando índices do relatório...",
    )
    conn.executescript("""
    CREATE INDEX IF NOT EXISTS idx_rel_mov_cp ON rel_movimentos_cp(cod_inc_cp);
    CREATE INDEX IF NOT EXISTS idx_rel_mov_rubr ON rel_movimentos_cp(cod_rubr, ide_tab_rubr);
    CREATE INDEX IF NOT EXISTS idx_rel_mov_trab ON rel_movimentos_cp(per_apur, cpf, matricula);
    """)
    conn.commit()

    _recriar_tabela_derivada(conn, "rel_movimentos_s2299", """
    SELECT r.*
    FROM dados_remuneracoes_s2299 AS r
    CROSS JOIN eventos AS e ON e.arquivo=r.arquivo
    WHERE e.tipo='S-2299'
      AND e.ativo=1
      AND e.duplicado_logico_de IS NULL
      AND NOT EXISTS (
        SELECT 1 FROM dados_exclusoes AS x
        WHERE COALESCE(x.nrRecEvt,'')<>''
          AND x.nrRecEvt=CASE
            WHEN COALESCE(e.recibo_evento,'')<>'' THEN e.recibo_evento
            WHEN COALESCE(e.ind_retif,'')<>'2'
             AND COALESCE(r.nr_recibo_evento,'')<>COALESCE(e.recibo_referencia,'')
              THEN COALESCE(r.nr_recibo_evento,'')
            ELSE '' END
      )
    """, cb, 0.42, "Preparando movimentos S-2299 para reconciliação...")
    conn.executescript("""
    CREATE INDEX IF NOT EXISTS idx_rel_s2299_per ON rel_movimentos_s2299(per_apur);
    CREATE INDEX IF NOT EXISTS idx_rel_s2299_chave
      ON rel_movimentos_s2299(per_apur,nr_insc_estab,cod_lotacao,cod_categ);
    """)
    conn.commit()

    _recriar_tabela_derivada(conn, "rel_rubricas_cp_base", """
    SELECT cod_rubr, ide_tab_rubr, dsc_rubr, nat_rubr, tp_rubr, cod_inc_cp,
           ini_valid, fim_valid, origem_bloco_s1010, criterio_cruzamento_s1010,
           origem_validacao, nivel_confianca, status_auditoria,
           status_cp, considerado_cp, tipo_verba, carater_verba,
           SUM(CAST(vr_rubr AS REAL)) valor_total,
           COUNT(*) qtd_lancamentos,
           COUNT(DISTINCT cpf) qtd_cpfs,
           MIN(per_apur) primeira_competencia,
           MAX(per_apur) ultima_competencia
    FROM rel_movimentos_cp
    GROUP BY cod_rubr, ide_tab_rubr, dsc_rubr, nat_rubr, tp_rubr, cod_inc_cp,
             ini_valid, fim_valid, origem_bloco_s1010, criterio_cruzamento_s1010,
             origem_validacao, nivel_confianca, status_auditoria,
             status_cp, considerado_cp, tipo_verba, carater_verba
    """, cb, 0.48, "Consolidando rubricas e totais...")

    _recriar_tabela_derivada(conn, "rel_sem_s1010", """
    SELECT per_apur, cpf, matricula, cod_rubr, ide_tab_rubr, dsc_rubr,
           SUM(CAST(vr_rubr AS REAL)) valor_rubrica, COUNT(*) qtd_lancamentos
    FROM rel_movimentos_cp WHERE status_cp='Sem S-1010'
    GROUP BY per_apur, cpf, matricula, cod_rubr, ide_tab_rubr, dsc_rubr
    """, cb, 0.60, "Consolidando ocorrências sem S-1010...")

    _recriar_tabela_derivada(conn, "rel_s5001_resumo", """
    SELECT cpf, matricula, per_apur, cod_categ, tp_insc_estab, nr_insc_estab,
           cod_lotacao, tp_valor, SUM(CAST(valor AS REAL)) valor_s5001, COUNT(*) qtd_linhas_s5001
    FROM dados_bases_trabalhador
    WHERE origem_valor='infoBaseCS' AND COALESCE(nr_recibo_base,'') NOT IN (
        SELECT COALESCE(nrRecEvt,'') FROM dados_exclusoes WHERE COALESCE(nrRecEvt,'')<>''
    )
    GROUP BY cpf, matricula, per_apur, cod_categ, tp_insc_estab, nr_insc_estab, cod_lotacao, tp_valor
    """, cb, 0.70, "Consolidando bases oficiais S-5001...")

    _recriar_tabela_derivada(conn, "rel_base_trabalhador", """
    WITH teorica AS (
      SELECT per_apur, cpf, matricula, cod_categ, cod_lotacao,
             SUM(CAST(vr_rubr AS REAL)) total_s1200,
             SUM(CASE WHEN considerado_cp='Sim' THEN CAST(vr_rubr AS REAL) ELSE 0 END) total_incide_cp,
             SUM(CASE WHEN status_cp='Não incide CP' THEN CAST(vr_rubr AS REAL) ELSE 0 END) total_nao_incide_cp,
             COUNT(*) qtd_rubricas,
             SUM(CASE WHEN status_cp='Sem S-1010' THEN 1 ELSE 0 END) qtd_rubricas_sem_s1010
      FROM rel_movimentos_cp GROUP BY per_apur, cpf, matricula, cod_categ, cod_lotacao
    ), oficial AS (
      SELECT per_apur, cpf, matricula, cod_categ, cod_lotacao,
             SUM(valor_s5001) total_s5001_infoBaseCS
      FROM rel_s5001_resumo GROUP BY per_apur, cpf, matricula, cod_categ, cod_lotacao
    )
    SELECT t.*, COALESCE(o.total_s5001_infoBaseCS,0) total_s5001_infoBaseCS,
           t.total_incide_cp-COALESCE(o.total_s5001_infoBaseCS,0) diferenca_incide_cp_vs_s5001,
           CASE WHEN ABS(t.total_incide_cp-COALESCE(o.total_s5001_infoBaseCS,0))<=0.05 THEN 'OK' ELSE 'Revisar' END status_conferencia
    FROM teorica t LEFT JOIN oficial o USING(per_apur,cpf,matricula,cod_categ,cod_lotacao)
    UNION ALL
    SELECT o.per_apur,o.cpf,o.matricula,o.cod_categ,o.cod_lotacao,
           0,0,0,0,0,o.total_s5001_infoBaseCS,-o.total_s5001_infoBaseCS,
           CASE WHEN ABS(o.total_s5001_infoBaseCS)<=0.05 THEN 'OK' ELSE 'Revisar' END
    FROM oficial o LEFT JOIN teorica t USING(per_apur,cpf,matricula,cod_categ,cod_lotacao)
    WHERE t.cpf IS NULL
    """, cb, 0.82, "Consolidando bases por trabalhador...")

    _recriar_tabela_derivada(conn, "rel_reconciliacao_s5011", """
    WITH s1200 AS (
      SELECT per_apur,nr_insc_estab,cod_lotacao,cod_categ,
             SUM(CASE WHEN considerado_cp='Sim' THEN CAST(vr_rubr AS REAL) ELSE 0 END) base_s1200
      FROM rel_movimentos_cp
      GROUP BY per_apur,nr_insc_estab,cod_lotacao,cod_categ
    ), s2299 AS (
      SELECT per_apur,nr_insc_estab,cod_lotacao,cod_categ,
             SUM(CASE WHEN considerado_cp='Sim' THEN CAST(vr_rubr AS REAL) ELSE 0 END) base_s2299
      FROM rel_movimentos_s2299
      GROUP BY per_apur,nr_insc_estab,cod_lotacao,cod_categ
    ), oficial AS (
      SELECT per_apur,nr_insc_estab,cod_lotacao,cod_categ,
             SUM(CAST(vr_bc_cp AS REAL)) base_s5011
      FROM dados_bases_contribuicao
      GROUP BY per_apur,nr_insc_estab,cod_lotacao,cod_categ
    ), chaves AS (
      SELECT per_apur,nr_insc_estab,cod_lotacao,cod_categ FROM s1200
      UNION SELECT per_apur,nr_insc_estab,cod_lotacao,cod_categ FROM s2299
      UNION SELECT per_apur,nr_insc_estab,cod_lotacao,cod_categ FROM oficial
    )
    SELECT c.per_apur,c.nr_insc_estab,c.cod_lotacao,c.cod_categ,
           COALESCE(a.base_s1200,0) base_s1200,
           COALESCE(d.base_s2299,0) base_s2299,
           COALESCE(o.base_s5011,0) base_s5011,
           COALESCE(a.base_s1200,0)+COALESCE(d.base_s2299,0) base_teorica_total,
           COALESCE(a.base_s1200,0)-COALESCE(o.base_s5011,0) diferenca_s1200_s5011,
           COALESCE(a.base_s1200,0)+COALESCE(d.base_s2299,0)-COALESCE(o.base_s5011,0) diferenca_total_s5011,
           CASE
             WHEN o.per_apur IS NULL THEN 'Sem S-5011'
             WHEN ABS(COALESCE(a.base_s1200,0)+COALESCE(d.base_s2299,0)-COALESCE(o.base_s5011,0))<=0.05 THEN 'OK'
             ELSE 'Revisar'
           END status_reconciliacao
    FROM chaves c
    LEFT JOIN s1200 a USING(per_apur,nr_insc_estab,cod_lotacao,cod_categ)
    LEFT JOIN s2299 d USING(per_apur,nr_insc_estab,cod_lotacao,cod_categ)
    LEFT JOIN oficial o USING(per_apur,nr_insc_estab,cod_lotacao,cod_categ)
    """, cb, 0.90, "Reconciliando S-1200 e S-2299 com S-5011...")
    conn.commit()
    emitir_progresso(
        cb, "consolidacao", 0.95,
        "Consolidações dos relatórios concluídas.",
    )
    emitir_progresso(
        cb, "integridade", 0.0,
        "Criando controles de integridade do relatório...",
    )
    _recriar_tabela_derivada(conn, "rel_controle_integridade", """
    SELECT 'rel_movimentos_cp' tabela, COUNT(*) quantidade, COALESCE(SUM(CAST(vr_rubr AS REAL)),0) soma_valores FROM rel_movimentos_cp
    UNION ALL SELECT 'dados_bases_trabalhador', COUNT(*), COALESCE(SUM(CAST(valor AS REAL)),0) FROM dados_bases_trabalhador
    UNION ALL SELECT 'dados_bases_contribuicao', COUNT(*), COALESCE(SUM(CAST(vr_bc_cp AS REAL)),0) FROM dados_bases_contribuicao
    UNION ALL SELECT 'rel_movimentos_s2299', COUNT(*), COALESCE(SUM(CAST(vr_rubr AS REAL)),0) FROM rel_movimentos_s2299
    UNION ALL SELECT 'rel_reconciliacao_s5011', COUNT(*), COALESCE(SUM(CAST(diferenca_total_s5011 AS REAL)),0) FROM rel_reconciliacao_s5011
    UNION ALL SELECT 'eventos', COUNT(*), COALESCE(SUM(tamanho_bytes),0) FROM eventos
    """, cb, 0.98, "Atualizando os controles de integridade...")
    conn.commit()
    emitir_progresso(
        cb, "integridade", 1.0,
        "Controles de integridade concluídos.",
        "Consolidação SQLite concluída sem carregar bases gigantes na memória.",
    )


def reconstruir_consolidados_analiticos(
    conn: sqlite3.Connection,
    progress_callback: ProgressCallback | None = None,
) -> None:
    """Reconstrói somente os consolidados a partir das tabelas analíticas."""
    _criar_tabelas_relatorio(conn, progress_callback)


def _metricas_movimentos(conn: sqlite3.Connection) -> dict[str, float | int]:
    """Calcula todos os totais de movimentos em uma unica varredura."""
    row = conn.execute("""
        SELECT
            COUNT(*) total,
            COALESCE(SUM(CAST(vr_rubr AS REAL)), 0) valor_total,
            COALESCE(SUM(CASE WHEN considerado_cp='Sim'
                THEN CAST(vr_rubr AS REAL) ELSE 0 END), 0) valor_incide,
            COALESCE(SUM(CASE WHEN status_cp='Não incide CP'
                THEN CAST(vr_rubr AS REAL) ELSE 0 END), 0) valor_nao_incide,
            COALESCE(SUM(CASE WHEN status_cp='Incide CP' THEN 1 ELSE 0 END), 0) qtd_padrao,
            COALESCE(SUM(CASE WHEN status_cp IN ('Incide CP','Revisar codIncCP')
                THEN 1 ELSE 0 END), 0) qtd_prioritaria,
            COALESCE(SUM(CASE WHEN status_cp IN
                ('Incide CP','Não incide CP','Revisar codIncCP')
                THEN 1 ELSE 0 END), 0) qtd_classificados,
            COALESCE(SUM(CASE WHEN status_cp='Sem S-1010' THEN 1 ELSE 0 END), 0) qtd_sem_s1010
        FROM rel_movimentos_cp
    """).fetchone()
    nomes = (
        "total", "valor_total", "valor_incide", "valor_nao_incide",
        "qtd_padrao", "qtd_prioritaria", "qtd_classificados", "qtd_sem_s1010",
    )
    return dict(zip(nomes, row))


def carregar_pacote_resumido(db_path: str | Path, limite_previa: int = 5_000) -> dict[str, object]:
    conn = sqlite3.connect(str(db_path))
    try:
        rub = pd.read_sql_query("SELECT * FROM rel_rubricas_cp_base ORDER BY status_cp, carater_verba, valor_total DESC", conn)
        if not rub.empty:
            rub["prioridade_revisao"] = rub.apply(_prioridade_rubrica, axis=1)
            rub["observacao"] = rub.apply(_observacao_rubrica, axis=1)
        resumo_rows = []
        metricas = _metricas_movimentos(conn)
        def scalar(sql: str): return conn.execute(sql).fetchone()[0]
        def existe_tabela(nome: str) -> bool:
            return conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (nome,)
            ).fetchone() is not None
        qtd_s2299 = int(scalar("SELECT COUNT(*) FROM rel_movimentos_s2299")) if existe_tabela("rel_movimentos_s2299") else 0
        valor_s2299 = float(scalar("SELECT COALESCE(SUM(CAST(vr_rubr AS REAL)),0) FROM rel_movimentos_s2299")) if existe_tabela("rel_movimentos_s2299") else 0.0
        qtd_reconciliacao_revisar = int(scalar("SELECT COUNT(*) FROM rel_reconciliacao_s5011 WHERE status_reconciliacao='Revisar'")) if existe_tabela("rel_reconciliacao_s5011") else 0
        resumo_rows.extend([
            {"indicador":"Rubricas únicas no S-1200","valor":len(rub)},
            {"indicador":"Rubricas com incidência CP","valor":int((rub.get("status_cp",pd.Series(dtype=str))=="Incide CP").sum())},
            {"indicador":"Rubricas sem incidência CP","valor":int((rub.get("status_cp",pd.Series(dtype=str))=="Não incide CP").sum())},
            {"indicador":"Rubricas sem S-1010","valor":int((rub.get("status_cp",pd.Series(dtype=str))=="Sem S-1010").sum())},
            {"indicador":"Valor total S-1200","valor":float(metricas["valor_total"])},
            {"indicador":"Valor com incidência CP","valor":float(metricas["valor_incide"])},
            {"indicador":"Valor sem incidência CP","valor":float(metricas["valor_nao_incide"])},
            {"indicador":"Linhas S-5001 detalhadas","valor":int(scalar("SELECT COUNT(*) FROM dados_bases_trabalhador"))},
            {"indicador":"Movimentos S-2299 segregados","valor":qtd_s2299},
            {"indicador":"Valor S-2299 segregado","valor":valor_s2299},
            {"indicador":"Reconciliações S-5011 para revisar","valor":qtd_reconciliacao_revisar},
        ])
        reconciliacao = (
            pd.read_sql_query(
                f"SELECT * FROM rel_reconciliacao_s5011 "
                f"ORDER BY status_reconciliacao DESC,ABS(diferenca_total_s5011) DESC "
                f"LIMIT {int(limite_previa)}",
                conn,
            )
            if existe_tabela("rel_reconciliacao_s5011")
            else pd.DataFrame()
        )
        return {
            "resumo_visual": pd.DataFrame(resumo_rows),
            "rubricas_cp": rub,
            "movimentos_cp": pd.read_sql_query(f"SELECT * FROM rel_movimentos_cp LIMIT {int(limite_previa)}", conn),
            "base_trabalhador": pd.read_sql_query(f"SELECT * FROM rel_base_trabalhador ORDER BY status_conferencia DESC LIMIT {int(limite_previa)}", conn),
            "sem_cadastro": pd.read_sql_query(f"SELECT * FROM rel_sem_s1010 ORDER BY valor_rubrica DESC LIMIT {int(limite_previa)}", conn),
            "s5001_resumo": pd.read_sql_query(f"SELECT * FROM rel_s5001_resumo LIMIT {int(limite_previa)}", conn),
            "controle_integridade": pd.read_sql_query("SELECT * FROM rel_controle_integridade", conn),
            "reconciliacao_s5011": reconciliacao,
            "total_movimentos_s2299": qtd_s2299,
            "total_movimentos_cp": int(metricas["total"]),
            "total_movimentos_cp_padrao": int(metricas["qtd_padrao"]),
            "total_movimentos_analise_prioritaria": int(metricas["qtd_prioritaria"]),
            "total_movimentos_classificados": int(metricas["qtd_classificados"]),
            "total_movimentos_sem_s1010": int(metricas["qtd_sem_s1010"]),
        }
    finally:
        conn.close()


def _nome_aba(nome: str, parte: int | None = None) -> str:
    if parte is None:
        return nome[:31]
    suf = f"_{parte}"
    return f"{nome[:31-len(suf)]}{suf}"


def _escrever_dataframe(ws, df: pd.DataFrame) -> tuple[int, float]:
    ws.append(list(df.columns))
    soma = 0.0
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    return len(df), soma


def _exportar_query(wb: Workbook, conn: sqlite3.Connection, nome: str, query: str, params: tuple = (), cb=None, inicio=0.0, fim=1.0) -> dict:
    total = int(conn.execute(f"SELECT COUNT(*) FROM ({query}) q", params).fetchone()[0])
    partes = max(1, math.ceil(total / MAX_DADOS_ABA))
    exportados = 0
    for parte in range(1, partes + 1):
        ws = wb.create_sheet(_nome_aba(nome, parte if partes > 1 else None))
        offset = (parte - 1) * MAX_DADOS_ABA
        limite = min(MAX_DADOS_ABA, total - offset)
        primeira = True
        cur = conn.execute(query + " LIMIT ? OFFSET ?", params + (limite, offset))
        cols = [d[0] for d in cur.description]
        ws.append(cols)
        while True:
            rows = cur.fetchmany(CHUNK_EXPORT)
            if not rows:
                break
            for row in rows:
                ws.append(list(row))
            exportados += len(rows)
            _progresso(cb, inicio + (fim-inicio) * exportados/max(total,1), f"Exportando {nome}: {exportados:,} de {total:,}".replace(",", "."))
    return {"aba": nome, "quantidade_sqlite": total, "quantidade_exportada": exportados, "partes": partes, "status": "OK" if total == exportados else "REVISAR"}




def _filtro_sql_movimentos_exportacao(modo: str) -> str:
    """Retorna o WHERE da aba 03; Sem S-1010 permanece no relatório 05."""
    if modo == "sem_s1010":
        return " WHERE status_cp='Sem S-1010'"
    if modo == "incidencia_cp_padrao":
        return " WHERE status_cp='Incide CP'"
    if modo == "analise_prioritaria":
        return " WHERE status_cp IN ('Incide CP','Revisar codIncCP')"
    if modo == "classificados":
        return " WHERE status_cp IN ('Incide CP','Não incide CP','Revisar codIncCP')"
    return ""


def contar_movimentos_exportacao_sqlite(
    db_path: str | Path,
    modo_exportacao_movimentos_cp: str,
    chaves_rubricas: list[str] | set[str] | None = None,
) -> int:
    """Conta a exportação com os mesmos filtros que serão usados no workbook."""
    filtro = combinar_filtros_sql(
        _filtro_sql_movimentos_exportacao(modo_exportacao_movimentos_cp),
        filtro_sql_chaves_rubricas(chaves_rubricas),
    )
    conn = sqlite3.connect(
        f"file:{Path(db_path).resolve().as_posix()}?mode=ro", uri=True, timeout=120
    )
    try:
        conn.execute("PRAGMA query_only = ON")
        conn.execute("PRAGMA busy_timeout = 120000")
        return int(
            conn.execute("SELECT COUNT(*) FROM rel_movimentos_cp" + filtro).fetchone()[0]
        )
    finally:
        conn.close()

def gerar_excel_saida_sqlite(
    db_path: str | Path,
    caminho_saida: str | Path,
    df_empresa: pd.DataFrame,
    df_resumo_visual: pd.DataFrame,
    df_rubricas_cp: pd.DataFrame,
    df_levantamento: pd.DataFrame | None = None,
    modo_exportacao_movimentos_cp: str = "todos",
    chaves_rubricas: list[str] | set[str] | None = None,
    progress_callback: ProgressCallback | None = None,
    gerar_manifesto: bool = False,
) -> str:
    """Gera um único relatório V10 em streaming e valida antes da entrega."""
    from modules.processador_zip import preparar_workspace_para_relatorios

    preparar_workspace_para_relatorios(
        db_path,
        progress_callback=progress_callback,
    )
    destino = Path(caminho_saida)
    destino.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(
        f"file:{Path(db_path).resolve().as_posix()}?mode=ro",
        uri=True,
        timeout=120,
    )
    conn.execute("PRAGMA query_only = ON")
    conn.execute("PRAGMA busy_timeout = 120000")
    conn.execute("PRAGMA temp_store = FILE")
    conn.execute("PRAGMA cache_size = -262144")
    try:
        condicao_rubricas = filtro_sql_chaves_rubricas(chaves_rubricas)
        filtro = combinar_filtros_sql(
            _filtro_sql_movimentos_exportacao(modo_exportacao_movimentos_cp),
            condicao_rubricas,
        )
        filtro_rubricas = combinar_filtros_sql("", condicao_rubricas)
        metricas = _metricas_movimentos(conn)
        totais_controle = {
            str(tabela): int(quantidade)
            for tabela, quantidade in conn.execute(
                "SELECT tabela,quantidade FROM rel_controle_integridade"
            )
        }
        totais_modo = {
            "todos": int(metricas["total"]),
            "incidencia_cp_padrao": int(metricas["qtd_padrao"]),
            "analise_prioritaria": int(metricas["qtd_prioritaria"]),
            "classificados": int(metricas["qtd_classificados"]),
        }
        fontes = [
            FontePlanilha("00_empresa", dataframe=df_empresa),
            FontePlanilha("01_resumo", dataframe=df_resumo_visual),
            FontePlanilha("02_rubricas_cp", dataframe=df_rubricas_cp),
            FontePlanilha(
                "03_movimentos_cp",
                query="SELECT * FROM rel_movimentos_cp" + filtro,
                total_esperado=(
                    contar_movimentos_exportacao_sqlite(
                        db_path, modo_exportacao_movimentos_cp, chaves_rubricas
                    )
                    if chaves_rubricas is not None
                    else totais_modo.get(
                        modo_exportacao_movimentos_cp, int(metricas["total"])
                    )
                ),
            ),
            FontePlanilha(
                "04_base_trabalhador",
                query="SELECT * FROM rel_base_trabalhador ORDER BY status_conferencia DESC, per_apur, cpf, matricula",
            ),
            FontePlanilha(
                "05_sem_s1010",
                query="SELECT * FROM rel_sem_s1010" + filtro_rubricas
                + " ORDER BY per_apur, valor_rubrica DESC",
            ),
            FontePlanilha(
                "05_busca_recibos",
                query="SELECT cod_rubr,ide_tab_rubr,MAX(dsc_rubr) dsc_rubr,MIN(per_apur) primeira_competencia,MAX(per_apur) ultima_competencia,COUNT(DISTINCT per_apur) qtd_competencias,SUM(qtd_lancamentos) qtd_lancamentos,COUNT(DISTINCT cpf) qtd_cpfs,SUM(valor_rubrica) valor_total,'Baixar recibo S-1010 da rubrica e aplicar na complementação do relatório.' acao_recomendada FROM rel_sem_s1010"
                + filtro_rubricas + " GROUP BY cod_rubr,ide_tab_rubr",
            ),
            FontePlanilha("06_s5001_tpvalor", query="SELECT * FROM rel_s5001_resumo"),
            FontePlanilha(
                "07_levantamento",
                dataframe=df_levantamento
                if df_levantamento is not None
                else pd.DataFrame(),
            ),
            FontePlanilha(
                "08_reconciliacao_s5011",
                query="SELECT * FROM rel_reconciliacao_s5011 "
                "ORDER BY status_reconciliacao DESC,ABS(diferenca_total_s5011) DESC",
                total_esperado=totais_controle.get("rel_reconciliacao_s5011"),
            ),
            FontePlanilha(
                "09_versoes",
                query="SELECT chave,valor FROM meta WHERE chave IN "
                "('versao_engine','versao_schema_sqlite','versao_parser',"
                "'versao_consistencia_recibo_s1200') ORDER BY chave",
            ),
            FontePlanilha(
                "apoio_s1010", query="SELECT * FROM dados_rubricas" + filtro_rubricas
            ),
            FontePlanilha(
                "apoio_s1200",
                query="SELECT * FROM rel_movimentos_cp" + filtro_rubricas,
                total_esperado=(
                    contar_movimentos_exportacao_sqlite(db_path, "todos", chaves_rubricas)
                    if chaves_rubricas is not None
                    else int(metricas["total"])
                ),
            ),
            FontePlanilha(
                "apoio_s2299",
                query="SELECT * FROM rel_movimentos_s2299" + filtro_rubricas,
                total_esperado=(
                    None
                    if chaves_rubricas is not None
                    else totais_controle.get("rel_movimentos_s2299")
                ),
            ),
            FontePlanilha(
                "apoio_s5001", query="SELECT * FROM dados_bases_trabalhador",
                total_esperado=totais_controle.get("dados_bases_trabalhador"),
            ),
            FontePlanilha(
                "apoio_s5011", query="SELECT * FROM dados_bases_contribuicao",
                total_esperado=totais_controle.get("dados_bases_contribuicao"),
            ),
            FontePlanilha("apoio_s3000", query="SELECT * FROM dados_exclusoes"),
            FontePlanilha(
                "checagem_layout",
                query="SELECT c.tipo,c.quantidade xml_localizados,CASE WHEN c.tipo IN ('S-1200','S-2299','S-5001','S-5011') THEN (SELECT COUNT(*) FROM eventos e WHERE e.tipo=c.tipo AND e.processado_segunda=1) ELSE c.quantidade END xml_parseados,0 nao_parseados FROM contagem_eventos c",
            ),
            FontePlanilha(
                "inventario",
                query="SELECT arquivo,tipo,tamanho_bytes,CASE WHEN tipo IN ('S-1000','S-1005','S-1010','S-1020','S-1200','S-2200','S-2299','S-3000','S-5001','S-5011') THEN 1 ELSE 0 END parseado,CASE envelope_recibo WHEN 1 THEN 'Sim' ELSE 'Não' END envelope_recibo FROM eventos ORDER BY id",
                total_esperado=totais_controle.get("eventos"),
            ),
            FontePlanilha(
                "erros_xml", query="SELECT arquivo,erro FROM erros ORDER BY id"
            ),
        ]
        banco_df = pd.read_sql_query("SELECT * FROM rel_controle_integridade", conn)
        fontes.append(FontePlanilha("controle_banco", dataframe=banco_df))
        resultado = gerar_workbook(
            destino,
            fontes,
            conexao=conn,
            progress_callback=progress_callback,
            gerar_manifesto=gerar_manifesto,
        )
        return resultado.caminho
    finally:
        conn.close()


# ============================================================
# V9.4.1 - Exportação segmentada com TEMP isolada e persistente
# ============================================================
MAX_LINHAS_ARQUIVO_SEGMENTADO = 200_000


def _definir_temp_ativo(pasta_temp: Path) -> Path:
    """Define uma pasta TEMP válida para SQLite, tempfile e openpyxl."""
    pasta_temp = Path(pasta_temp).expanduser().resolve()
    pasta_temp.mkdir(parents=True, exist_ok=True)
    caminho = str(pasta_temp)
    os.environ["TMP"] = caminho
    os.environ["TEMP"] = caminho
    os.environ["TMPDIR"] = caminho
    tempfile.tempdir = caminho
    return pasta_temp


def _preparar_temp_sqlite(pasta_saida: Path) -> Path:
    """Cria uma TEMP persistente para consultas SQLite durante toda a exportação."""
    pasta_temp_sqlite = _definir_temp_ativo(pasta_saida / ".sqlite_temp")
    teste = pasta_temp_sqlite / "teste_gravacao.tmp"
    try:
        teste.write_text("ok", encoding="utf-8")
        teste.unlink(missing_ok=True)
    except OSError as exc:
        raise OSError(
            f"A pasta temporária do SQLite não permite gravação: {pasta_temp_sqlite}"
        ) from exc
    return pasta_temp_sqlite


def _limpar_temp_openpyxl(pasta_temp: Path) -> None:
    """Remove somente os temporários criados pela exportação atual."""
    if pasta_temp.exists():
        shutil.rmtree(pasta_temp, ignore_errors=True)


def _novo_workbook_em_temp(pasta_temp: Path) -> Workbook:
    _definir_temp_ativo(pasta_temp)
    return Workbook(write_only=True)


def _salvar_workbook_isolado(
    wb: Workbook,
    destino: Path,
    pasta_temp_openpyxl: Path,
    pasta_temp_sqlite: Path,
) -> None:
    """Salva o XLSX, remove sua TEMP e restaura a TEMP persistente do SQLite."""
    try:
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)
    finally:
        _limpar_temp_openpyxl(pasta_temp_openpyxl)
        _definir_temp_ativo(pasta_temp_sqlite)


def _exportar_query_segmentada(
    conn: sqlite3.Connection,
    pasta_saida: Path,
    pasta_temp_sqlite: Path,
    prefixo_arquivo: str,
    nome_aba: str,
    query: str,
    params: tuple = (),
    linhas_por_arquivo: int = MAX_LINHAS_ARQUIVO_SEGMENTADO,
    progress_callback: ProgressCallback | None = None,
    inicio: float = 0.0,
    fim: float = 1.0,
) -> list[dict]:
    _definir_temp_ativo(pasta_temp_sqlite)
    try:
        resultado = conn.execute(
            f"SELECT COUNT(*) FROM ({query}) AS q",
            params,
        ).fetchone()
        total = int(resultado[0] if resultado else 0)
    except sqlite3.OperationalError as exc:
        raise sqlite3.OperationalError(
            f"Falha ao consultar o conjunto '{prefixo_arquivo}'. "
            f"Banco: {conn.execute('PRAGMA database_list').fetchone()[2]}. "
            f"TEMP ativa: {pasta_temp_sqlite}. Erro original: {exc}"
        ) from exc

    partes = max(1, math.ceil(total / max(1, linhas_por_arquivo)))
    exportados = 0
    controles: list[dict] = []

    for parte in range(1, partes + 1):
        offset = (parte - 1) * linhas_por_arquivo
        limite = min(linhas_por_arquivo, max(total - offset, 0))
        nome_arq = (
            f"{prefixo_arquivo}_parte_{parte:03d}.xlsx"
            if partes > 1
            else f"{prefixo_arquivo}.xlsx"
        )
        destino = pasta_saida / nome_arq
        pasta_temp_openpyxl = (
            pasta_saida / ".openpyxl_temp" / f"{prefixo_arquivo}_{parte:03d}"
        )

        wb = _novo_workbook_em_temp(pasta_temp_openpyxl)
        ws = wb.create_sheet(_nome_aba(nome_aba))

        _definir_temp_ativo(pasta_temp_sqlite)
        cur = conn.execute(
            query + " LIMIT ? OFFSET ?",
            params + (limite, offset),
        )
        cols = [d[0] for d in cur.description]
        ws.append(cols)

        qtd_parte = 0
        while True:
            rows = cur.fetchmany(CHUNK_EXPORT)
            if not rows:
                break
            for row in rows:
                ws.append(list(row))
            qtd_parte += len(rows)
            exportados += len(rows)
            _progresso(
                progress_callback,
                inicio + (fim - inicio) * exportados / max(total, 1),
                f"Exportando {prefixo_arquivo}: {exportados:,} de {total:,}".replace(",", "."),
            )

        cur.close()
        _salvar_workbook_isolado(
            wb,
            destino,
            pasta_temp_openpyxl,
            pasta_temp_sqlite,
        )
        controles.append(
            {
                "arquivo": nome_arq,
                "conjunto": prefixo_arquivo,
                "parte": parte,
                "quantidade_sqlite": total,
                "quantidade_exportada_parte": qtd_parte,
                "offset_inicial": offset,
                "status": "OK",
            }
        )

    return controles


def gerar_pacote_excel_saida_sqlite(
    db_path: str | Path,
    pasta_saida: str | Path,
    df_empresa: pd.DataFrame,
    df_resumo_visual: pd.DataFrame,
    df_rubricas_cp: pd.DataFrame,
    df_levantamento: pd.DataFrame | None = None,
    modo_exportacao_movimentos_cp: str = "todos",
    chaves_rubricas: list[str] | set[str] | None = None,
    linhas_por_arquivo: int = MAX_LINHAS_ARQUIVO_SEGMENTADO,
    progress_callback: ProgressCallback | None = None,
    gerar_manifesto: bool = False,
) -> dict:
    """Compatibilidade V9.4: entrega um único workbook consolidado V10."""
    saida_compat = Path(pasta_saida).expanduser().resolve()
    caminho_unico = saida_compat / "relatorio_incidencia_cp_esocial_v10.xlsx"
    caminho = gerar_excel_saida_sqlite(
        db_path=db_path,
        caminho_saida=caminho_unico,
        df_empresa=df_empresa,
        df_resumo_visual=df_resumo_visual,
        df_rubricas_cp=df_rubricas_cp,
        df_levantamento=df_levantamento,
        modo_exportacao_movimentos_cp=modo_exportacao_movimentos_cp,
        chaves_rubricas=chaves_rubricas,
        progress_callback=progress_callback,
        gerar_manifesto=gerar_manifesto,
    )
    manifesto = Path(caminho).with_name(f"{Path(caminho).stem}_manifesto.csv")
    return {
        "pasta_saida": str(Path(caminho).parent),
        "arquivos": [caminho],
        "manifesto": str(manifesto) if manifesto.exists() else "",
        "quantidade_arquivos": 1,
    }

    # Implementacao segmentada V9.4 mantida abaixo apenas como referencia
    # temporaria de compatibilidade; nao e mais alcancavel nem usada pela V9.5.
    saida = Path(str(pasta_saida).strip().strip('"').strip("'")).expanduser().resolve()
    saida.mkdir(parents=True, exist_ok=True)

    banco = Path(str(db_path).strip().strip('"').strip("'")).expanduser().resolve()
    if not banco.exists():
        raise FileNotFoundError(f"Banco SQLite não localizado: {banco}")
    if not banco.is_file():
        raise ValueError(f"O caminho informado não é um arquivo SQLite: {banco}")

    pasta_temp_sqlite = _preparar_temp_sqlite(saida)

    conn = sqlite3.connect(
        f"file:{banco.as_posix()}?mode=ro",
        uri=True,
        timeout=120,
    )
    conn.execute("PRAGMA query_only = ON")
    conn.execute("PRAGMA busy_timeout = 120000")
    conn.execute("PRAGMA temp_store = FILE")
    conn.execute("PRAGMA cache_size = -262144")

    controles: list[dict] = []
    try:
        _progresso(progress_callback, 0.01, "Preparando relatório segmentado...")

        principal = saida / "00_relatorio_principal.xlsx"
        temp_principal = saida / ".openpyxl_temp" / "principal"
        wb = _novo_workbook_em_temp(temp_principal)
        for nome, df in [
            ("00_empresa", df_empresa),
            ("01_resumo", df_resumo_visual),
            ("02_rubricas_cp", df_rubricas_cp),
            (
                "07_levantamento",
                df_levantamento if df_levantamento is not None else pd.DataFrame(),
            ),
        ]:
            ws = wb.create_sheet(nome)
            _escrever_dataframe(ws, df if df is not None else pd.DataFrame())

        _salvar_workbook_isolado(
            wb,
            principal,
            temp_principal,
            pasta_temp_sqlite,
        )
        controles.append(
            {
                "arquivo": principal.name,
                "conjunto": "principal",
                "parte": 1,
                "quantidade_exportada_parte": 0,
                "status": "OK",
            }
        )

        filtro = _filtro_sql_movimentos_exportacao(modo_exportacao_movimentos_cp)
        consultas = [
            ("03_movimentos_cp", "03_movimentos_cp", "SELECT * FROM rel_movimentos_cp" + filtro, 0.05, 0.48),
            ("04_base_trabalhador", "04_base_trabalhador", "SELECT * FROM rel_base_trabalhador ORDER BY status_conferencia DESC, per_apur, cpf, matricula", 0.48, 0.60),
            ("05_sem_s1010", "05_sem_s1010", "SELECT * FROM rel_sem_s1010 ORDER BY per_apur, valor_rubrica DESC", 0.60, 0.66),
            ("05_busca_recibos", "05_busca_recibos", "SELECT cod_rubr,ide_tab_rubr,MAX(dsc_rubr) dsc_rubr,MIN(per_apur) primeira_competencia,MAX(per_apur) ultima_competencia,COUNT(DISTINCT per_apur) qtd_competencias,SUM(qtd_lancamentos) qtd_lancamentos,COUNT(DISTINCT cpf) qtd_cpfs,SUM(valor_rubrica) valor_total,'Baixar recibo S-1010 da rubrica e aplicar na complementação do relatório.' acao_recomendada FROM rel_sem_s1010 GROUP BY cod_rubr,ide_tab_rubr", 0.66, 0.69),
            ("06_s5001_tpvalor", "06_s5001_tpvalor", "SELECT * FROM rel_s5001_resumo", 0.69, 0.73),
            ("apoio_s1010", "apoio_s1010", "SELECT * FROM dados_rubricas", 0.73, 0.77),
            ("apoio_s1200", "apoio_s1200", "SELECT * FROM rel_movimentos_cp", 0.77, 0.86),
            ("apoio_s5001", "apoio_s5001", "SELECT * FROM dados_bases_trabalhador", 0.86, 0.91),
            ("apoio_s5011", "apoio_s5011", "SELECT * FROM dados_bases_contribuicao", 0.91, 0.94),
            ("apoio_s3000", "apoio_s3000", "SELECT * FROM dados_exclusoes", 0.94, 0.955),
            ("checagem_layout", "checagem_layout", "SELECT c.tipo,c.quantidade xml_localizados,CASE WHEN c.tipo IN ('S-1200','S-2299','S-5001','S-5011') THEN (SELECT COUNT(*) FROM eventos e WHERE e.tipo=c.tipo AND e.processado_segunda=1) ELSE c.quantidade END xml_parseados,0 nao_parseados FROM contagem_eventos c", 0.955, 0.965),
            ("inventario", "inventario", "SELECT arquivo,tipo,tamanho_bytes,CASE WHEN tipo IN ('S-1000','S-1005','S-1010','S-1020','S-1200','S-2200','S-2299','S-3000','S-5001','S-5011') THEN 1 ELSE 0 END parseado,CASE envelope_recibo WHEN 1 THEN 'Sim' ELSE 'Não' END envelope_recibo FROM eventos ORDER BY id", 0.965, 0.985),
            ("erros_xml", "erros_xml", "SELECT arquivo,erro FROM erros ORDER BY id", 0.985, 0.992),
        ]

        for prefixo, aba, sql, ini, fim in consultas:
            controles.extend(
                _exportar_query_segmentada(
                    conn=conn,
                    pasta_saida=saida,
                    pasta_temp_sqlite=pasta_temp_sqlite,
                    prefixo_arquivo=prefixo,
                    nome_aba=aba,
                    query=sql,
                    linhas_por_arquivo=linhas_por_arquivo,
                    progress_callback=progress_callback,
                    inicio=ini,
                    fim=fim,
                )
            )

        manifesto = pd.DataFrame(controles)
        manifesto_path = saida / "manifesto_exportacao.csv"
        manifesto.to_csv(
            manifesto_path,
            index=False,
            sep=";",
            encoding="utf-8-sig",
        )

        _definir_temp_ativo(pasta_temp_sqlite)
        banco_ctrl = pd.read_sql_query(
            "SELECT * FROM rel_controle_integridade",
            conn,
        )
        banco_ctrl.to_csv(
            saida / "controle_integridade_banco.csv",
            index=False,
            sep=";",
            encoding="utf-8-sig",
        )

        _limpar_temp_openpyxl(saida / ".openpyxl_temp")
        _progresso(
            progress_callback,
            1.0,
            "Relatório segmentado concluído sem perda de dados.",
        )
        return {
            "pasta_saida": str(saida),
            "arquivos": [str(saida / c["arquivo"]) for c in controles],
            "manifesto": str(manifesto_path),
            "quantidade_arquivos": len(controles),
        }
    finally:
        conn.close()
        _limpar_temp_openpyxl(saida / ".openpyxl_temp")

