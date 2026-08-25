import sqlite3
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.data_source import SQLiteDataSource, WorkspaceContext
from modules.levantamento_sqlite import (
    FiltrosLevantamento,
    competencias_no_intervalo,
    consultar_levantamento,
    consultar_rubricas,
    gerar_excel_levantamento_sqlite,
    obter_opcoes_filtros,
)


COLUNAS_MOVIMENTOS = [
    "cod_rubr", "ide_tab_rubr", "dsc_rubr", "nat_rubr", "cod_inc_cp",
    "status_cp", "carater_verba", "tipo_verba", "per_apur", "cpf", "vr_rubr",
]


class LevantamentoSQLiteTest(unittest.TestCase):
    def test_competencias_no_intervalo_inclusivo(self):
        competencias = ["2026-01", "2026-02", "2026-03", "2026-04"]
        self.assertEqual(
            competencias_no_intervalo(competencias, "2026-02", "2026-03"),
            ["2026-02", "2026-03"],
        )
        self.assertEqual(
            competencias_no_intervalo(competencias, "2026-02", "2026-02"),
            ["2026-02"],
        )

    def test_competencias_no_intervalo_rejeita_ordem_invalida(self):
        with self.assertRaisesRegex(ValueError, "inicial"):
            competencias_no_intervalo(
                ["2026-01", "2026-02"], "2026-02", "2026-01"
            )

    def _workspace(self, pasta: str) -> tuple[Path, pd.DataFrame]:
        workspace = Path(pasta) / "esocial_v3_teste"
        workspace.mkdir()
        db = workspace / "processamento.db"
        dados = pd.DataFrame(
            [
                ["100", "1", "Salário", "1000", "11", "Incide CP", "Mensal", "Remuneratória", "2026-01", "111", 100.0],
                ["100", "1", "Salário", "1000", "11", "Incide CP", "Mensal", "Remuneratória", "2026-01", "222", 200.0],
                ["100", "1", "Salário", "1000", "11", "Incide CP", "Mensal", "Remuneratória", "2026-02", "111", -10.0],
                ["200", "1", "Bônus", "1200", "00", "Não incide CP", "Eventual", "Remuneratória", "2026-01", "111", 50.0],
                ["300", "1", "Sem cadastro", "", "", "Sem S-1010", "Revisar", "Outros", "2026-02", "333", 75.0],
            ],
            columns=COLUNAS_MOVIMENTOS,
        )
        conn = sqlite3.connect(db)
        try:
            conn.execute("CREATE TABLE meta(chave TEXT PRIMARY KEY, valor TEXT)")
            conn.execute("INSERT INTO meta VALUES('status','concluido')")
            conn.execute("CREATE TABLE dados_empresa(nome_empresa TEXT,cnpj_empregador TEXT)")
            conn.execute("INSERT INTO dados_empresa VALUES('Empresa Teste','12345678000199')")
            dados.to_sql("rel_movimentos_cp", conn, index=False)
            conn.execute("""
                CREATE TABLE rel_rubricas_cp_base AS
                SELECT cod_rubr,ide_tab_rubr,dsc_rubr,cod_inc_cp,status_cp,
                       carater_verba,tipo_verba,
                       SUM(CAST(vr_rubr AS REAL)) valor_total,
                       COUNT(*) qtd_lancamentos,
                       COUNT(DISTINCT cpf) qtd_cpfs,
                       MIN(per_apur) primeira_competencia,
                       MAX(per_apur) ultima_competencia
                FROM rel_movimentos_cp
                GROUP BY cod_rubr,ide_tab_rubr,dsc_rubr,cod_inc_cp,status_cp,
                         carater_verba,tipo_verba
            """)
            conn.commit()
        finally:
            conn.close()
        return workspace, dados

    def test_contexto_aceita_pasta_e_banco(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, _ = self._workspace(pasta)
            por_pasta = WorkspaceContext.from_path(workspace)
            por_banco = WorkspaceContext.from_path(workspace / "processamento.db")
            self.assertEqual(por_pasta.db_path, por_banco.db_path)
            self.assertEqual(por_pasta.status, "concluido")

    def test_contexto_rejeita_workspace_interrompido(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, _ = self._workspace(pasta)
            conn = sqlite3.connect(workspace / "processamento.db")
            try:
                conn.execute("UPDATE meta SET valor='interrompido' WHERE chave='status'")
                conn.commit()
            finally:
                conn.close()
            with self.assertRaisesRegex(ValueError, "não está concluído"):
                WorkspaceContext.from_path(workspace)

    def test_filtros_e_resumos_equivalem_ao_pandas(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, dados = self._workspace(pasta)
            fonte = SQLiteDataSource(WorkspaceContext.from_path(workspace))
            filtros = FiltrosLevantamento(
                status_cp="Incide CP", apenas_positivos=True
            )
            rubricas = consultar_rubricas(fonte, filtros)
            resultado = consultar_levantamento(
                fonte, filtros, ["100||1"], aliquota=20.0
            )
            esperado = dados[
                dados["status_cp"].eq("Incide CP")
                & (pd.to_numeric(dados["vr_rubr"]) > 0)
            ]
            self.assertEqual(len(rubricas), 1)
            self.assertAlmostEqual(resultado.total, esperado["vr_rubr"].sum())
            self.assertAlmostEqual(resultado.cpp, esperado["vr_rubr"].sum() * 0.20)
            self.assertEqual(resultado.qtd_movimentos, len(esperado))
            self.assertEqual(resultado.qtd_cpfs, esperado["cpf"].nunique())

    def test_opcoes_e_sem_s1010(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, _ = self._workspace(pasta)
            fonte = SQLiteDataSource(WorkspaceContext.from_path(workspace))
            opcoes = obter_opcoes_filtros(fonte)
            self.assertIn("Incide CP", opcoes["status_cp"])
            self.assertIn("Sem S-1010", opcoes["cod_inc_cp"])
            self.assertEqual(opcoes["competencias"], ["2026-01", "2026-02"])
            rubricas = consultar_rubricas(
                fonte,
                FiltrosLevantamento(cod_inc_cp="Sem S-1010", apenas_positivos=False),
            )
            self.assertEqual(rubricas["cod_rubr"].tolist(), ["300"])

    def test_catalogo_consolidado_nao_altera_filtro_final_de_positivos(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, _ = self._workspace(pasta)
            fonte = SQLiteDataSource(WorkspaceContext.from_path(workspace))
            filtros = FiltrosLevantamento(
                status_cp="Incide CP", apenas_positivos=True
            )
            rubricas = consultar_rubricas(fonte, filtros)
            resultado = consultar_levantamento(
                fonte, filtros, ["100||1"], aliquota=20.0
            )
            self.assertEqual(rubricas["cod_rubr"].tolist(), ["100"])
            self.assertAlmostEqual(float(rubricas.iloc[0]["valor_total"]), 290.0)
            self.assertAlmostEqual(resultado.total, 300.0)
            self.assertEqual(resultado.qtd_movimentos, 2)

    def test_catalogo_consolidado_filtra_por_sobreposicao_do_periodo(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, _ = self._workspace(pasta)
            fonte = SQLiteDataSource(WorkspaceContext.from_path(workspace))

            janeiro = consultar_rubricas(
                fonte,
                FiltrosLevantamento(
                    competencias=("2026-01",), apenas_positivos=False
                ),
            )
            marco = consultar_rubricas(
                fonte,
                FiltrosLevantamento(
                    competencias=("2026-03",), apenas_positivos=False
                ),
            )

            self.assertEqual(set(janeiro["cod_rubr"]), {"100", "200"})
            self.assertTrue(marco.empty)

    def test_exportacao_detalhada_usa_sqlite(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, _ = self._workspace(pasta)
            fonte = SQLiteDataSource(WorkspaceContext.from_path(workspace))
            filtros = FiltrosLevantamento(status_cp="Incide CP")
            resultado = consultar_levantamento(fonte, filtros, ["100||1"], 20.0)
            destino = Path(pasta) / "levantamento.xlsx"
            gerar_excel_levantamento_sqlite(
                fonte,
                destino,
                filtros,
                ["100||1"],
                resultado,
                pd.DataFrame([{"nome_empresa": "Empresa Teste"}]),
                pd.DataFrame([{"Indicador": "Teste", "Valor": "OK"}]),
                True,
            )
            workbook = load_workbook(destino, read_only=True)
            try:
                self.assertIn("03_movimentos", workbook.sheetnames)
                self.assertIn("controle_integridade", workbook.sheetnames)
            finally:
                workbook.close()

    def test_muitas_rubricas_nao_dependem_do_limite_de_parametros(self):
        with tempfile.TemporaryDirectory() as pasta:
            workspace, _ = self._workspace(pasta)
            conn = sqlite3.connect(workspace / "processamento.db")
            try:
                linhas = [
                    (
                        f"R{i:04d}", "1", f"Rubrica {i}", "1000", "11",
                        "Incide CP", "Mensal", "Remuneratória", "2026-03",
                        f"{i:011d}", 1.0,
                    )
                    for i in range(1_100)
                ]
                conn.executemany(
                    "INSERT INTO rel_movimentos_cp VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                    linhas,
                )
                conn.commit()
            finally:
                conn.close()
            fonte = SQLiteDataSource(WorkspaceContext.from_path(workspace))
            chaves = [f"R{i:04d}||1" for i in range(1_100)]
            resultado = consultar_levantamento(
                fonte,
                FiltrosLevantamento(status_cp="Incide CP"),
                chaves,
                20.0,
            )
            self.assertEqual(resultado.qtd_movimentos, 1_100)
            self.assertEqual(resultado.qtd_rubricas, 1_100)


if __name__ == "__main__":
    unittest.main()
