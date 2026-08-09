import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from modules.excel_builder import FontePlanilha, gerar_workbook, validar_workbook


class ExcelBuilderTest(unittest.TestCase):
    def test_validacao_rapida_nao_percorre_linhas(self):
        with tempfile.TemporaryDirectory() as pasta:
            destino = Path(pasta) / "rapido.xlsx"
            resultado = gerar_workbook(
                destino,
                [FontePlanilha("dados", dataframe=pd.DataFrame({"id": [1, 2]}))],
            )
            with patch(
                "openpyxl.worksheet._read_only.ReadOnlyWorksheet.iter_rows",
                side_effect=AssertionError("iter_rows nao deveria ser chamado"),
            ):
                abas = validar_workbook(destino, resultado.controles)
            self.assertIn("dados", abas)

    def test_validacao_completa_opcional_confere_linhas(self):
        with tempfile.TemporaryDirectory() as pasta:
            destino = Path(pasta) / "completo.xlsx"
            resultado = gerar_workbook(
                destino,
                [FontePlanilha("dados", dataframe=pd.DataFrame({"id": [1, 2]}))],
            )
            abas = validar_workbook(
                destino,
                resultado.controles,
                validacao_completa=True,
            )
            self.assertIn("dados", abas)

    def test_dataframe_pequeno_e_sem_s1010(self):
        with tempfile.TemporaryDirectory() as pasta:
            destino = Path(pasta) / "relatorio.xlsx"
            resultado = gerar_workbook(
                destino,
                [
                    FontePlanilha(
                        "00_empresa",
                        dataframe=pd.DataFrame(
                            [{"nome_empresa": "Empresa Teste", "cnpj": "123"}]
                        ),
                    ),
                    FontePlanilha(
                        "05_sem_s1010",
                        dataframe=pd.DataFrame(columns=["cod_rubr", "valor"]),
                    ),
                    FontePlanilha(
                        "apoio_s1010",
                        dataframe=pd.DataFrame(
                            [{"cod_rubr": "1000", "cod_inc_cp": "11"}]
                        ),
                    ),
                ],
            )
            self.assertEqual(resultado.manifesto, None)
            self.assertTrue(destino.exists())
            workbook = load_workbook(destino, read_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "00_empresa",
                        "05_sem_s1010",
                        "apoio_s1010",
                        "controle_integridade",
                    ],
                )
            finally:
                workbook.close()

    def test_volume_medio_em_varios_chunks(self):
        with tempfile.TemporaryDirectory() as pasta:
            conexao = sqlite3.connect(":memory:")
            try:
                conexao.execute("CREATE TABLE movimentos(id INTEGER)")
                conexao.executemany(
                    "INSERT INTO movimentos VALUES(?)", ((i,) for i in range(2501))
                )
                resultado = gerar_workbook(
                    Path(pasta) / "medio.xlsx",
                    [
                        FontePlanilha(
                            "03_movimentos_cp",
                            query="SELECT * FROM movimentos ORDER BY id",
                        )
                    ],
                    conexao=conexao,
                    chunk=100,
                )
            finally:
                conexao.close()
            self.assertEqual(len(resultado.controles), 1)
            self.assertEqual(resultado.controles[0].linhas_exportadas, 2501)

    def test_query_incremental_divide_abas_e_manifesto_opcional(self):
        with tempfile.TemporaryDirectory() as pasta:
            conexao = sqlite3.connect(":memory:")
            try:
                conexao.execute("CREATE TABLE movimentos(id INTEGER, status_cp TEXT)")
                conexao.executemany(
                    "INSERT INTO movimentos VALUES(?,?)",
                    [(i, "Incide CP") for i in range(5)],
                )
                destino = Path(pasta) / "grande.xlsx"
                resultado = gerar_workbook(
                    destino,
                    [
                        FontePlanilha(
                            "03_movimentos_cp",
                            query="SELECT * FROM movimentos ORDER BY id",
                        )
                    ],
                    conexao=conexao,
                    gerar_manifesto=True,
                    max_dados_aba=2,
                    chunk=1,
                )
            finally:
                conexao.close()

            self.assertEqual(
                [c.linhas_exportadas for c in resultado.controles], [2, 2, 1]
            )
            self.assertTrue(Path(resultado.manifesto).exists())
            workbook = load_workbook(destino, read_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "03_movimentos_cp_1",
                        "03_movimentos_cp_2",
                        "03_movimentos_cp_3",
                        "controle_integridade",
                    ],
                )
            finally:
                workbook.close()

    def test_total_esperado_evitar_contagem_redundante(self):
        with tempfile.TemporaryDirectory() as pasta:
            conexao = sqlite3.connect(":memory:")
            try:
                conexao.execute("CREATE TABLE movimentos(id INTEGER)")
                conexao.executemany(
                    "INSERT INTO movimentos VALUES(?)", [(1,), (2,), (3,)]
                )
                with patch(
                    "modules.excel_builder._total_query",
                    side_effect=AssertionError("COUNT redundante"),
                ):
                    resultado = gerar_workbook(
                        Path(pasta) / "total_conhecido.xlsx",
                        [FontePlanilha(
                            "movimentos",
                            query="SELECT * FROM movimentos ORDER BY id",
                            total_esperado=3,
                        )],
                        conexao=conexao,
                    )
            finally:
                conexao.close()
            self.assertEqual(resultado.controles[0].linhas_exportadas, 3)


if __name__ == "__main__":
    unittest.main()
