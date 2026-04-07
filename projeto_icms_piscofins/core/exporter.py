from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _sheet_safe_name(name: str) -> str:
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in invalid:
        name = name.replace(ch, "-")
    return name[:31]


def ajustar_largura(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    ws = writer.sheets[sheet_name]

    if df.empty and len(df.columns) == 0:
        ws.column_dimensions["A"].width = 20
        return

    for i, col in enumerate(df.columns, 1):
        tamanho = len(str(col))
        if not df.empty:
            serie = df[col].astype(str).fillna("")
            try:
                tamanho = max(tamanho, int(serie.map(len).max()))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(i)].width = min(tamanho + 2, 40)


def estilizar_cabecalho(writer: pd.ExcelWriter, sheet_name: str) -> None:
    ws = writer.sheets[sheet_name]

    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def aplicar_cor_status(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    if df.empty or "Status da Análise" not in df.columns:
        return

    ws = writer.sheets[sheet_name]
    idx_status = list(df.columns).index("Status da Análise") + 1

    fill_verde = PatternFill(fill_type="solid", fgColor="E2F0D9")
    fill_vermelho = PatternFill(fill_type="solid", fgColor="FCE4D6")
    fill_amarelo = PatternFill(fill_type="solid", fgColor="FFF2CC")
    fill_cinza = PatternFill(fill_type="solid", fgColor="E7E6E6")

    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=idx_status).value

        if valor == "Exclusão identificada":
            fill = fill_verde
        elif valor == "Sem indício de exclusão":
            fill = fill_vermelho
        elif valor == "Divergente / Revisar":
            fill = fill_amarelo
        elif valor == "Sem dados suficientes":
            fill = fill_cinza
        else:
            fill = None

        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def criar_resumo(report: pd.DataFrame, resumo: dict[str, Any]) -> pd.DataFrame:
    total = int(resumo.get("itens_analisados", len(report)))
    exclusao = int(resumo.get("itens_exclusao_identificada", 0))
    sem_indicio = int(resumo.get("itens_sem_indicio", 0))
    divergente = int(resumo.get("itens_divergente_revisar", 0))
    sem_dados = int(resumo.get("itens_sem_dados", 0))
    sem_match = int(resumo.get("itens_sem_match", 0))

    base_icms = int(
        resumo.get(
            "itens_com_base_icms_final",
            int((pd.to_numeric(report.get("Base de ICMS Final", 0), errors="coerce").fillna(0) != 0).sum())
            if "Base de ICMS Final" in report.columns
            else 0,
        )
    )
    valor_icms = int(
        resumo.get(
            "itens_com_valor_icms_final",
            int((pd.to_numeric(report.get("Valor de ICMS Final", 0), errors="coerce").fillna(0) != 0).sum())
            if "Valor de ICMS Final" in report.columns
            else 0,
        )
    )

    credito_total = float(resumo.get("credito_total_estimado", 0.0))

    join_sim = int((report["Cruzou com ICMS/IPI"].astype(str) == "Sim").sum()) if "Cruzou com ICMS/IPI" in report.columns else 0
    join_nao = int((report["Cruzou com ICMS/IPI"].astype(str) == "Não").sum()) if "Cruzou com ICMS/IPI" in report.columns else 0
    itens_st = int((report["Possui ST"].astype(str) == "Sim").sum()) if "Possui ST" in report.columns else 0

    dados = [
        ["RESULTADO DA ANÁLISE", ""],
        ["Itens analisados", total],
        ["Exclusão identificada", exclusao],
        ["Sem indício de exclusão", sem_indicio],
        ["Divergente / Revisar", divergente],
        ["Sem dados suficientes", sem_dados],
        ["% com exclusão", exclusao / total if total else 0],
        ["% sem exclusão", sem_indicio / total if total else 0],
        ["Crédito total estimado", credito_total],
        ["", ""],
        ["QUALIDADE DOS DADOS", ""],
        ["Linhas com Base de ICMS Final", base_icms],
        ["Linhas com Valor de ICMS Final", valor_icms],
        ["Join Sim", join_sim],
        ["Join Não", join_nao],
        ["Itens com ICMS-ST", itens_st],
        ["Itens sem cruzamento", sem_match],
    ]

    return pd.DataFrame(dados, columns=["Indicador", "Valor"])


def simplificar(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=[
            "Empresa",
            "CNPJ",
            "Mês",
            "Ano",
            "Número da Nota",
            "Série",
            "Item",
            "Código do Produto",
            "Descrição",
            "CFOP",
            "Operação",
            "Base de ICMS Final",
            "Valor de ICMS Final",
            "Base de PIS Informada",
            "Base de COFINS Informada",
            "Diferença ICMS x PIS",
            "Diferença ICMS x COFINS",
            "Diferença bate com ICMS? PIS",
            "Diferença bate com ICMS? COFINS",
            "Status da Análise",
            "Motivo",
            "Crédito PIS Estimado",
            "Crédito COFINS Estimado",
            "Crédito Total Estimado",
        ])

    colunas = [
        "Empresa",
        "CNPJ",
        "Mês",
        "Ano",
        "Número da Nota",
        "Série",
        "Item",
        "Código do Produto",
        "Descrição",
        "CFOP",
        "Operação",
        "Base de ICMS Final",
        "Valor de ICMS Final",
        "Base de PIS Informada",
        "Base de COFINS Informada",
        "Diferença ICMS x PIS",
        "Diferença ICMS x COFINS",
        "Diferença bate com ICMS? PIS",
        "Diferença bate com ICMS? COFINS",
        "Status da Análise",
        "Motivo",
        "Crédito PIS Estimado",
        "Crédito COFINS Estimado",
        "Crédito Total Estimado",
    ]
    colunas_existentes = [c for c in colunas if c in df.columns]
    return df[colunas_existentes].copy()


def preparar_relatorio_completo(df: pd.DataFrame) -> pd.DataFrame:
    prioridade = [
        "Empresa",
        "CNPJ",
        "Mês",
        "Ano",
        "Chave",
        "Número da Nota",
        "Série",
        "Item",
        "Código do Produto",
        "Descrição",
        "CFOP",
        "Operação",
        "Valor do Item",
        "Base de ICMS Final",
        "Valor de ICMS Final",
        "Base de PIS Informada",
        "Base de COFINS Informada",
        "Diferença ICMS x PIS",
        "Diferença ICMS x COFINS",
        "Diferença bate com ICMS? PIS",
        "Diferença bate com ICMS? COFINS",
        "Status da Análise",
        "Motivo",
        "Crédito PIS Estimado",
        "Crédito COFINS Estimado",
        "Crédito Total Estimado",
        "Base de ICMS no PIS",
        "Valor de ICMS no PIS",
        "Base de ICMS no ICMS/IPI",
        "Valor de ICMS no ICMS/IPI",
        "Base de ICMS ST",
        "Valor de ICMS ST",
        "Origem da Base de ICMS",
        "Origem do Valor de ICMS",
        "Possui ST",
        "Cruzou com ICMS/IPI",
        "Documento de Entrada",
    ]

    cols1 = [c for c in prioridade if c in df.columns]
    cols2 = [c for c in df.columns if c not in cols1]
    out = df[cols1 + cols2].copy()

    if "Status da Análise" in out.columns:
        ordem = {
            "Sem indício de exclusão": 0,
            "Divergente / Revisar": 1,
            "Sem dados suficientes": 2,
            "Exclusão identificada": 3,
        }
        out["_ordem_status"] = out["Status da Análise"].map(ordem).fillna(9)
        sort_cols = ["_ordem_status"]
        if "Ano" in out.columns:
            sort_cols.append("Ano")
        if "Mês" in out.columns:
            sort_cols.append("Mês")
        if "Número da Nota" in out.columns:
            sort_cols.append("Número da Nota")
        if "Item" in out.columns:
            sort_cols.append("Item")
        out = out.sort_values(sort_cols, kind="stable").drop(columns="_ordem_status")

    return out


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> str:
    safe_name = _sheet_safe_name(sheet_name)

    if df is None or (df.empty and len(df.columns) == 0):
        df = pd.DataFrame({"Mensagem": ["Nenhum registro encontrado para esta aba."]})
    elif df.empty:
        df = df.copy()

    df.to_excel(writer, sheet_name=safe_name, index=False)
    ajustar_largura(writer, df, safe_name)
    estilizar_cabecalho(writer, safe_name)
    aplicar_cor_status(writer, df, safe_name)
    return safe_name


def export_report_to_bytes(report: pd.DataFrame, resumo: dict[str, Any]) -> bytes:
    report_export = report.copy()

    colunas_numericas = [
        "Valor do Item",
        "Base de ICMS no PIS",
        "Valor de ICMS no PIS",
        "Base de ICMS no ICMS/IPI",
        "Valor de ICMS no ICMS/IPI",
        "Base de ICMS Final",
        "Valor de ICMS Final",
        "Base de ICMS ST",
        "Valor de ICMS ST",
        "Base de PIS Informada",
        "Base de COFINS Informada",
        "Diferença ICMS x PIS",
        "Diferença ICMS x COFINS",
        "Crédito PIS Estimado",
        "Crédito COFINS Estimado",
        "Crédito Total Estimado",
    ]

    for col in colunas_numericas:
        if col in report_export.columns:
            report_export[col] = pd.to_numeric(report_export[col], errors="coerce")

    resumo_df = criar_resumo(report_export, resumo)

    if "Status da Análise" in report_export.columns:
        exclusao_df = simplificar(report_export[report_export["Status da Análise"] == "Exclusão identificada"].copy())
        sem_indicio_df = simplificar(report_export[report_export["Status da Análise"] == "Sem indício de exclusão"].copy())
        divergente_df = simplificar(report_export[report_export["Status da Análise"] == "Divergente / Revisar"].copy())
        sem_dados_df = simplificar(report_export[report_export["Status da Análise"] == "Sem dados suficientes"].copy())
    else:
        exclusao_df = simplificar(pd.DataFrame())
        sem_indicio_df = simplificar(pd.DataFrame())
        divergente_df = simplificar(pd.DataFrame())
        sem_dados_df = simplificar(pd.DataFrame())

    sem_match_df = simplificar(
        report_export[report_export["Cruzou com ICMS/IPI"] == "Não"].copy()
    ) if "Cruzou com ICMS/IPI" in report_export.columns else simplificar(pd.DataFrame())

    st_df = simplificar(
        report_export[report_export["Possui ST"] == "Sim"].copy()
    ) if "Possui ST" in report_export.columns else simplificar(pd.DataFrame())

    completo_df = preparar_relatorio_completo(report_export)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_sheet(writer, resumo_df, "Resumo Executivo")
        _write_sheet(writer, exclusao_df, "Exclusao Identificada")
        _write_sheet(writer, sem_indicio_df, "Sem Indicio")
        _write_sheet(writer, divergente_df, "Divergente Revisar")
        _write_sheet(writer, sem_dados_df, "Sem Dados")
        _write_sheet(writer, sem_match_df, "Sem Cruzamento")
        _write_sheet(writer, st_df, "Itens com ST")
        _write_sheet(writer, completo_df, "Relatorio Completo")

    output.seek(0)
    return output.getvalue()