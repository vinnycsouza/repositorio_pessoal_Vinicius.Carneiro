from __future__ import annotations

import io
import tempfile
import sqlite3
from pathlib import Path
from typing import Dict, Optional, Set, Tuple

import pandas as pd
from openpyxl import Workbook

from .layout import CAB_K300, CAB_K050
from .aggregate import montar_pivot_dtcomp_por_rubrica


def _norm_dt_comp_mmaaaa(x: str) -> str:
    return (str(x) or "").strip().zfill(6)


def _ord_dt_comp_mmaaaa(x: str) -> int:
    s = _norm_dt_comp_mmaaaa(x)
    if len(s) == 6 and s.isdigit():
        mm = int(s[:2])
        aaaa = int(s[2:])
        if 1 <= mm <= 12:
            return aaaa * 100 + mm
    return 99999999


def _write_df_to_sheet(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def _write_txt_event_to_sheet(ws, path_txt: Path, header: list[str], filter_fn=None):
    ws.append(header)
    with path_txt.open("r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            linha = linha.rstrip("\n")
            partes = linha.split("|")
            partes = partes[: len(header)]
            if len(partes) < len(header):
                partes += [""] * (len(header) - len(partes))
            if filter_fn and not filter_fn(partes):
                continue
            ws.append(partes)


def gerar_excel_interno(
    path_k300: Path,
    path_k150: Optional[Path],
    path_k050: Optional[Path],
    selected_codigos: Set[str],
    allowed_ind_rubr: Set[str],
    allowed_ind_base_ps: Set[str],
    df_rubricas: pd.DataFrame,
    aplicar_regra_terco_ferias: bool = False,
    rubricas_terco_ferias: Optional[Set[str]] = None,
) -> bytes:
    selected_codigos = set(map(str, selected_codigos))
    allowed_ind_rubr = set(map(str, allowed_ind_rubr))
    allowed_ind_base_ps = set(map(str, allowed_ind_base_ps))
    rubricas_terco_ferias = set(map(str, rubricas_terco_ferias or set()))

    cutoff_ord = 202009

    # mapa código -> descrição
    desc_map: Dict[str, str] = {}
    if df_rubricas is not None and not df_rubricas.empty:
        for _, r in df_rubricas.iterrows():
            desc_map[str(r.get("COD_RUBRICA", "")).strip()] = str(r.get("DESC_RUBRICA", "")).strip()

    wb = Workbook(write_only=True)

    # ----------------------------
    # 1) K300_FILTRADO (ordenado)
    # ----------------------------
    ws_k300 = wb.create_sheet(title="K300_FILTRADO")
    ws_k300.append(CAB_K300)

    def k300_keep(partes: list[str]) -> bool:
        cod = (partes[6] or "").strip()
        ind_r = (partes[8] or "").strip()
        ind_ps = (partes[10] or "").strip()
        dt_comp = _norm_dt_comp_mmaaaa(partes[5])

        if cod not in selected_codigos:
            return False
        if allowed_ind_rubr and ind_r not in allowed_ind_rubr:
            return False
        if allowed_ind_base_ps and ind_ps not in allowed_ind_base_ps:
            return False

        if aplicar_regra_terco_ferias and cod in rubricas_terco_ferias:
            if _ord_dt_comp_mmaaaa(dt_comp) > cutoff_ord:
                return False

        return True

    # Ordenação “por disco” (sem RAM): separa por DT_COMP em tempfiles e depois escreve em ordem
    tmp_dir = Path(tempfile.mkdtemp(prefix="k300_ord_"))
    buckets: Dict[str, Path] = {}

    with path_k300.open("r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            linha = linha.rstrip("\n")
            partes = linha.split("|")
            partes = partes[: len(CAB_K300)]
            if len(partes) < len(CAB_K300):
                partes += [""] * (len(CAB_K300) - len(partes))

            if not k300_keep(partes):
                continue

            dt_comp = _norm_dt_comp_mmaaaa(partes[5])
            if dt_comp not in buckets:
                buckets[dt_comp] = tmp_dir / f"{dt_comp}.txt"

            with buckets[dt_comp].open("a", encoding="utf-8", newline="\n") as out:
                out.write("|".join(partes) + "\n")

    # escreve os buckets em ordem cronológica
    for dt_comp in sorted(buckets.keys(), key=_ord_dt_comp_mmaaaa):
        p = buckets[dt_comp]
        with p.open("r", encoding="utf-8", errors="ignore") as bf:
            for linha in bf:
                linha = linha.rstrip("\n")
                partes = linha.split("|")
                partes = partes[: len(CAB_K300)]
                if len(partes) < len(CAB_K300):
                    partes += [""] * (len(CAB_K300) - len(partes))
                ws_k300.append(partes)

    # ----------------------------
    # 2) RESUMO_DT_COMP (pivot)
    # ----------------------------
    df_pivot = montar_pivot_dtcomp_por_rubrica(
        path_k300=path_k300,
        selected_codigos=selected_codigos,
        allowed_ind_rubr=allowed_ind_rubr,
        allowed_ind_base_ps=allowed_ind_base_ps,
        desc_map=desc_map,
        aplicar_regra_terco_ferias=aplicar_regra_terco_ferias,
        rubricas_terco_ferias=rubricas_terco_ferias,
    )
    ws_resumo = wb.create_sheet(title="RESUMO_DT_COMP")
    _write_df_to_sheet(ws_resumo, df_pivot)

    # ---------------------------------------------
    # 3) K300_PIVOT_TRAB (rubricas em colunas)
    # ---------------------------------------------
    # Para não estourar RAM em arquivo grande, uso SQLite temporário para agregação.
    db_path = tmp_dir / "k300_pivot.db"
    conn = sqlite3.connect(db_path.as_posix())
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS acc (
            REG TEXT,
            CNPJ_CEI TEXT,
            IND_FL TEXT,
            COD_LTC TEXT,
            COD_REG_TRAB TEXT,
            DT_COMP TEXT,
            COD_RUBR TEXT,
            TOTAL REAL,
            PRIMARY KEY (REG, CNPJ_CEI, IND_FL, COD_LTC, COD_REG_TRAB, DT_COMP, COD_RUBR)
        )
        """
    )
    conn.commit()

    def _parse_decimal_ptbr_to_float(v: str) -> float:
        v = (v or "").strip()
        if not v:
            return 0.0
        v = v.replace(".", "").replace(",", ".")
        try:
            return float(v)
        except Exception:
            return 0.0

    # agrega
    with path_k300.open("r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            linha = linha.rstrip("\n")
            partes = linha.split("|")
            partes = partes[: len(CAB_K300)]
            if len(partes) < len(CAB_K300):
                partes += [""] * (len(CAB_K300) - len(partes))

            if not k300_keep(partes):
                continue

            reg = (partes[0] or "").strip()
            cnpj = (partes[1] or "").strip()
            ind_fl = (partes[2] or "").strip()
            cod_ltc = (partes[3] or "").strip()
            cod_reg_trab = (partes[4] or "").strip()
            dt_comp = _norm_dt_comp_mmaaaa(partes[5])
            cod_rubr = (partes[6] or "").strip()
            valor = _parse_decimal_ptbr_to_float(partes[7])

            cur.execute(
                """
                INSERT INTO acc (REG, CNPJ_CEI, IND_FL, COD_LTC, COD_REG_TRAB, DT_COMP, COD_RUBR, TOTAL)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(REG, CNPJ_CEI, IND_FL, COD_LTC, COD_REG_TRAB, DT_COMP, COD_RUBR)
                DO UPDATE SET TOTAL = TOTAL + excluded.TOTAL
                """,
                (reg, cnpj, ind_fl, cod_ltc, cod_reg_trab, dt_comp, cod_rubr, valor),
            )
    conn.commit()

    # rubricas em colunas (ordem por código)
    rubricas_cols = sorted(selected_codigos, key=lambda x: (pd.to_numeric(str(x), errors="coerce"), str(x)))

    # cabeçalho da aba pivot
    ws_pivot_trab = wb.create_sheet(title="K300_PIVOT_TRAB")
    header = ["REG", "CNPJ/CEI", "IND_FL", "COD_LTC", "COD_REG_TRAB", "DT_COMP"] + [
        f"{cod} - {desc_map.get(cod, '')}".strip(" -")[:250] for cod in rubricas_cols
    ]
    ws_pivot_trab.append(header)

    # linhas distintas (ordenadas por DT_COMP cronológico + chaves)
    cur.execute(
        """
        SELECT DISTINCT REG, CNPJ_CEI, IND_FL, COD_LTC, COD_REG_TRAB, DT_COMP
        FROM acc
        """
    )
    keys = cur.fetchall()
    keys.sort(key=lambda r: (_ord_dt_comp_mmaaaa(r[5]), r[0], r[1], r[2], r[3], r[4]))

    # para cada chave, buscar valores por rubrica
    for (reg, cnpj, ind_fl, cod_ltc, cod_reg_trab, dt_comp) in keys:
        cur.execute(
            """
            SELECT COD_RUBR, TOTAL
            FROM acc
            WHERE REG=? AND CNPJ_CEI=? AND IND_FL=? AND COD_LTC=? AND COD_REG_TRAB=? AND DT_COMP=?
            """,
            (reg, cnpj, ind_fl, cod_ltc, cod_reg_trab, dt_comp),
        )
        got = dict(cur.fetchall())
        row = [reg, cnpj, ind_fl, cod_ltc, cod_reg_trab, dt_comp] + [float(got.get(c, 0.0)) for c in rubricas_cols]
        ws_pivot_trab.append(row)

    conn.close()

    # ----------------------------
    # 4) K150 selecionadas
    # ----------------------------
    ws_k150 = wb.create_sheet(title="K150_SELECIONADAS")
    ws_k150.append(["COD_RUBRICA", "DESC_RUBRICA"])
    if df_rubricas is not None and not df_rubricas.empty and selected_codigos:
        df_sel = df_rubricas.copy()
        df_sel["COD_RUBRICA"] = df_sel["COD_RUBRICA"].astype(str).str.strip()
        df_sel = df_sel[df_sel["COD_RUBRICA"].isin(selected_codigos)].drop_duplicates()
        df_sel["_COD_NUM"] = pd.to_numeric(df_sel["COD_RUBRICA"], errors="coerce")
        df_sel = df_sel.sort_values(by=["_COD_NUM", "COD_RUBRICA"], kind="stable", na_position="last").drop(columns=["_COD_NUM"])
        for _, r in df_sel.iterrows():
            ws_k150.append([str(r["COD_RUBRICA"]), str(r.get("DESC_RUBRICA", ""))])

    # ----------------------------
    # 5) K050 completo
    # ----------------------------
    if path_k050 and path_k050.exists():
        ws_k050 = wb.create_sheet(title="K050_TRABALHADORES")
        _write_txt_event_to_sheet(ws_k050, path_k050, CAB_K050, filter_fn=None)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()