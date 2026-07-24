from __future__ import annotations

from pathlib import Path
import csv
from typing import Iterable

from openpyxl import load_workbook


CABECALHOS = {"codigo", "código", "codrubr", "cod_rubr", "rubrica", "codigo_rubrica"}


def _normalizar(valor: object) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _unicos_preservando_ordem(valores: Iterable[object]) -> list[str]:
    vistos: set[str] = set()
    saida: list[str] = []
    for valor in valores:
        codigo = _normalizar(valor)
        if not codigo:
            continue
        if codigo.lower() in CABECALHOS:
            continue
        chave = codigo.upper()
        if chave not in vistos:
            vistos.add(chave)
            saida.append(codigo)
    return saida


def ler_codigos(caminho: str | Path) -> list[str]:
    arquivo = Path(caminho)
    extensao = arquivo.suffix.lower()

    if extensao in {".txt"}:
        return _unicos_preservando_ordem(
            linha.strip() for linha in arquivo.read_text(encoding="utf-8-sig", errors="ignore").splitlines()
        )

    if extensao == ".csv":
        valores: list[str] = []
        with arquivo.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
            amostra = f.read(4096)
            f.seek(0)
            try:
                dialeto = csv.Sniffer().sniff(amostra, delimiters=";,|\t,")
            except csv.Error:
                dialeto = csv.excel
                dialeto.delimiter = ";"
            leitor = csv.reader(f, dialeto)
            for linha in leitor:
                if linha:
                    valores.append(linha[0])
        return _unicos_preservando_ordem(valores)

    if extensao in {".xlsx", ".xlsm"}:
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active
        valores = []
        for linha in ws.iter_rows(values_only=True):
            primeiro = next((v for v in linha if v not in (None, "")), None)
            if primeiro is not None:
                valores.append(primeiro)
        wb.close()
        return _unicos_preservando_ordem(valores)

    raise ValueError(f"Formato não suportado: {extensao}")
