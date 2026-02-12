import streamlit as st
import pandas as pd
import io
import tempfile
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook



# =========================
# Cabeçalho por evento MANAD
# =========================
def cabecalho_evento(codigo):
    cabecalhos = {
        "I200": ["REG", "DT_LCTO", "COD_CTA", "COD_CCUS", "COD_CP", "VL_DEB_CRED",
                 "IND_DEB_CRED", "NUM_ARQ", "NUM_LCTO", "IND_LCTO", "HIST_LCTO"],
        "K300": ["REG", "CNPJ/CEI", "IND_FL", "COD_LTC", "COD_REG_TRAB", "DT_COMP",
                 "COD_RUBR", "VLR_RUBR", "IND_RUBR", "IND_BASE_IRRF", "IND_BASE_PS"],
        "K250": ["REG", "CNPJ/CEI", "IND_FL", "COD_LTC", "COD_REG_TRAB", "DT_COMP",
                 "DT_PGTO", "COD_CBO", "COD_OCORR", "DESC_CARGO", "QTD_DEP_IR",
                 "QTD_DEP_SF", "VL_BASE_IRRF", "VL_BASE_PS"],
        "K150": ["REG", "CNPJ/CEI", "DT_INC_ALT", "COD_RUBRICA", "DESC_RUBRICA"],
        "I050": ["REG", "DT_INC_ALT", "IND_NAT", "IND_GRP_CTA", "NIVEL",
                 "COD_GRP_CTA", "COD_GRP_CTA_SUP", "NOME_GRP_CTA"],
    }
    return cabecalhos.get(codigo)


def extrair_codigo_evento(linha: str) -> str | None:
    linha = (linha or "").strip()
    if len(linha) < 4:
        return None
    return linha[:4]


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="MANAD Extrator", layout="wide")
st.title("📂 MANAD - Separação por Evento e Geração de Excel")

uploaded_file = st.file_uploader(
    "Selecione o arquivo MANAD (.txt ou .xlsx)",
    type=["txt", "xlsx"],
    key="upload_manad"
)

# =========================
# Estado da sessão
# =========================
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None

if "estatisticas" not in st.session_state:
    st.session_state.estatisticas = {}

if "eventos_encontrados" not in st.session_state:
    st.session_state.eventos_encontrados = []


# =========================
# Leitura do arquivo e Prévia
# =========================
if uploaded_file:
    st.success("Arquivo carregado com sucesso!")

    # Para TXT grande: vamos “spoolar” em disco por evento (sem estourar RAM)
    # Para XLSX: mantemos o seu comportamento (mas é mais pesado)
    is_txt = uploaded_file.name.lower().endswith(".txt")

    # Pasta temporária para armazenar arquivos por evento
    tmp_dir = Path(tempfile.mkdtemp(prefix="manad_"))
    st.caption(f"📌 Processamento usando pasta temporária: {tmp_dir}")

    # Vamos guardar os arquivos por evento (spool) e contar linhas por evento
    arquivos_evento = {}  # codigo -> Path
    contagem_linhas = defaultdict(int)

    # =========================================
    # 1) Spool: separar em arquivos temporários
    # =========================================
    status = st.empty()
    progresso = st.progress(0.0)

    if is_txt:
        status.info("Separando o TXT por evento (modo econômico de memória)...")

        # Cria handles sob demanda (pra não abrir 200 arquivos sem necessidade)
        handles = {}

        def get_handle(codigo: str):
            if codigo not in handles:
                p = tmp_dir / f"{codigo}.txt"
                arquivos_evento[codigo] = p
                handles[codigo] = p.open("w", encoding="utf-8", newline="\n")
            return handles[codigo]

        # Percorre linha a linha (não usa read())
        total_bytes = getattr(uploaded_file, "size", None) or 1
        bytes_lidos = 0

        for raw in uploaded_file:
            bytes_lidos += len(raw)
            # Atualiza progresso por bytes (aproximado)
            progresso.progress(min(bytes_lidos / total_bytes, 1.0))

            linha = raw.decode("latin1", errors="ignore").rstrip("\n")
            codigo = extrair_codigo_evento(linha)
            if not codigo:
                continue

            # grava a linha completa no arquivo do evento
            h = get_handle(codigo)
            h.write(linha + "\n")
            contagem_linhas[codigo] += 1

        # fecha todos
        for h in handles.values():
            h.close()

        eventos = sorted(arquivos_evento.keys())
        st.session_state.eventos_encontrados = eventos
        status.success(f"Eventos encontrados: {', '.join(eventos)}")

    else:
        # XLSX como entrada: lê e faz spool por evento (pode ser mais lento/pesado)
        status.info("Lendo XLSX e separando por evento...")
        xls = pd.ExcelFile(uploaded_file)
        handles = {}

        def get_handle(codigo: str):
            if codigo not in handles:
                p = tmp_dir / f"{codigo}.txt"
                arquivos_evento[codigo] = p
                handles[codigo] = p.open("w", encoding="utf-8", newline="\n")
            return handles[codigo]

        total_abas = len(xls.sheet_names) or 1
        for i, aba in enumerate(xls.sheet_names, start=1):
            progresso.progress(i / total_abas)
            status.text(f"Lendo aba: {aba} ({i}/{total_abas})")
            df_aba = pd.read_excel(xls, sheet_name=aba, header=None)

            # Espera-se que esteja na coluna 0, como você já faz
            for v in df_aba[0].dropna().astype(str).tolist():
                v = v.strip()
                codigo = extrair_codigo_evento(v)
                if not codigo:
                    continue
                h = get_handle(codigo)
                h.write(v + "\n")
                contagem_linhas[codigo] += 1

        for h in handles.values():
            h.close()

        eventos = sorted(arquivos_evento.keys())
        st.session_state.eventos_encontrados = eventos
        status.success(f"Eventos encontrados: {', '.join(eventos)}")

    progresso.empty()

    # =========================================
    # 2) Gerar Excel (streaming) mantendo formato
    # =========================================
    from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

# ...

if st.button("⚙️ Gerar arquivo Excel por evento", key="gerar_excel"):
    MAX_ROWS_EXCEL = 1_048_576
    MAX_DADOS_POR_ABA = MAX_ROWS_EXCEL - 1

    progresso2 = st.progress(0.0)
    status2 = st.empty()

    eventos = st.session_state.eventos_encontrados
    total_eventos = len(eventos) or 1
    estatisticas = {}

    wb = Workbook(write_only=True)

    for idx_evento, codigo in enumerate(eventos, start=1):
        progresso2.progress(idx_evento / total_eventos)
        status2.text(f"Gerando Excel: evento {codigo} ({idx_evento}/{total_eventos})")

        cabecalho = cabecalho_evento(codigo)
        if not cabecalho:
            continue

        path_evento = arquivos_evento.get(codigo)
        if not path_evento or not path_evento.exists():
            continue

        total_linhas = contagem_linhas.get(codigo, 0)
        if total_linhas <= 0:
            continue

        total_abas = (total_linhas // MAX_DADOS_POR_ABA) + (1 if total_linhas % MAX_DADOS_POR_ABA else 0)

        estatisticas[codigo] = {
            "Total de linhas": int(total_linhas),
            "Total de abas": int(total_abas),
        }

        # índices e controle de aba
        aba_idx = 1
        linhas_na_aba = 0

        # cria primeira aba
        nome_aba = codigo if total_abas == 1 else f"{codigo}_{aba_idx}"
        ws = wb.create_sheet(title=nome_aba[:31])
        ws.append(cabecalho)
        aba_idx += 1

        with path_evento.open("r", encoding="utf-8", errors="ignore") as f:
            for linha in f:
                linha = linha.rstrip("\n")
                partes = linha.split("|")

                partes = partes[:len(cabecalho)]
                if len(partes) < len(cabecalho):
                    partes += [""] * (len(cabecalho) - len(partes))

                ws.append(partes)
                linhas_na_aba += 1

                # bateu o limite da aba: cria a próxima (se ainda tiver mais dados)
                if linhas_na_aba >= MAX_DADOS_POR_ABA and aba_idx <= total_abas:
                    linhas_na_aba = 0
                    nome_aba = codigo if total_abas == 1 else f"{codigo}_{aba_idx}"
                    ws = wb.create_sheet(title=nome_aba[:31])
                    ws.append(cabecalho)
                    aba_idx += 1

    progresso2.empty()
    status2.success("✅ Excel gerado com sucesso!")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    st.session_state.excel_bytes = output.getvalue()
    st.session_state.estatisticas = estatisticas
    
    # =========================
    # Download
    # =========================
    if st.session_state.excel_bytes:
        st.download_button(
            label="📥 Baixar Excel com todos os eventos",
            data=st.session_state.excel_bytes,
            file_name="MANAD_Eventos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_manad"
        )

    # =========================
    # Estatísticas
    # =========================
    if st.session_state.estatisticas:
        st.subheader("📊 Estatísticas por evento")
        df_stats = pd.DataFrame.from_dict(
            st.session_state.estatisticas,
            orient="index"
        ).reset_index()
        df_stats.columns = ["Evento", "Total de linhas", "Total de abas"]
        st.dataframe(df_stats, use_container_width=True)
