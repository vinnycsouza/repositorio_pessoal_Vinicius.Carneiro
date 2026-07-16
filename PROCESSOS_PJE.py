from __future__ import annotations

import re
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# ============================================================
# CONFIGURAÇÕES
# ============================================================

URL_CONSULTA = "https://pje.tst.jus.br/consultaprocessual/"
ARQUIVO_SAIDA = Path("status_processos_tst.xlsx")
PASTA_DIAGNOSTICO = Path("diagnosticos_tst")

TEMPO_ESPERA = 30
INTERVALO_ENTRE_PROCESSOS = 2
EXECUTAR_EM_SEGUNDO_PLANO = False
SALVAR_DIAGNOSTICO_DE_SUCESSO = False


# ============================================================
# PROCESSOS
# ============================================================

processos = [
    "0000915-57.2024.5.22.0006",
    "0000805-83.2023.5.22.0106",
    "0001765-44.2024.5.07.0034",
    "0001091-48.2024.5.22.0002",
    "0000991-02.2023.5.07.0017",
    "0000031-78.2025.5.06.0122",
    "0000135-46.2024.5.07.0003",
    "0001379-10.2022.5.05.0561",
    "0000883-97.2024.5.20.0009",
    "0001290-36.2023.5.06.0201",
    "0000380-27.2022.5.06.0271",
    "0000244-98.2022.5.19.0002",
    "0000235-75.2024.5.22.0005",
    "0001004-06.2024.5.06.0013",
    "0000147-19.2025.5.19.0059",
    "0000085-03.2023.5.19.0009",
    "0000354-39.2024.5.22.0004",
    "0000049-68.2024.5.19.0059",
    "0000734-23.2024.5.06.0161",
    "0000541-96.2023.5.06.0143",
    "0001723-54.2023.5.05.0561",
    "0001323-71.2021.5.06.0144",
    "0000637-68.2022.5.06.0201",
    "0001646-94.2022.5.10.0802",
    "0000750-57.2025.5.19.0006",
    "0000935-08.2023.5.07.0004",
    "0000940-94.2024.5.07.0036",
    "0000274-46.2023.5.10.0812",
    "0001066-80.2024.5.07.0025",
    "0000235-61.2025.5.06.0401",
    "0000437-72.2024.5.06.0401",
    "0000809-76.2024.5.19.0007",
    "0000783-92.2024.5.13.0024",
    "0000826-68.2023.5.22.0006",
    "0000903-98.2022.5.19.0005",
    "0000498-36.2022.5.05.0463",
    "0000672-56.2017.5.19.0002",
    "0000684-80.2024.5.06.0101",
    "0001634-09.2016.5.05.0195",
    "0000236-37.2021.5.19.0009",
    "0000593-85.2024.5.20.0008",
    "0000441-56.2022.5.05.0030",
    "0000112-70.2015.5.05.0036",
    "0000470-88.2022.5.05.0133",
    "0000889-94.2024.5.06.0009",
    "0000850-62.2019.5.06.0142",
    "0000920-40.2022.5.05.0421",
    "0000459-27.2022.5.05.0661",
    "0000376-79.2024.5.06.0251",
    "0000063-66.2024.5.07.0033",
    "0000707-06.2022.5.06.0001",
    "0000630-15.2023.5.05.0122",
    "0000958-24.2022.5.06.0001",
    "0000359-69.2024.5.07.0007",
    "0000287-68.2024.5.22.0103",
    "0000750-31.2023.5.06.0413",
    "0000303-29.2021.5.06.0020",
    "0001473-31.2024.5.07.0011",
    "0000997-24.2023.5.19.0001",
    "0000354-39.2024.5.22.0004",
    "0001255-28.2024.5.06.0141",
    "0000658-14.2022.5.06.0017",
    "0000628-12.2023.5.05.0133",
    "0000277-96.2024.5.19.0009",
    "0000156-71.2023.5.06.0201",
    "0000921-31.2024.5.06.0161",
    "0000670-25.2022.5.05.0221",
    "0000012-24.2023.5.05.0493",
    "0000668-91.2023.5.05.0133",
    "0000915-74.2024.5.07.0006",
    "0000044-23.2024.5.06.0022",
    "0000388-32.2023.5.06.0412",
    "0000009-93.2022.5.05.0464",
    "0000300-73.2024.5.22.0004",
]


TERMOS_TRANSITO = [
    "trânsito em julgado",
    "certidão de trânsito em julgado",
    "transitado em julgado",
]

TERMOS_CAPTCHA = [
    "captcha",
    "não sou um robô",
    "nao sou um robo",
    "verificação de segurança",
    "verificacao de seguranca",
]

TERMOS_NAO_ENCONTRADO = [
    "processo não encontrado",
    "processo nao encontrado",
    "nenhum processo encontrado",
    "nenhum resultado encontrado",
    "não foram encontrados processos",
    "nao foram encontrados processos",
]


# ============================================================
# UTILIDADES
# ============================================================

def criar_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    if EXECUTAR_EM_SEGUNDO_PLANO:
        options.add_argument("--headless=new")
    options.add_argument("--start-maximized")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    return webdriver.Chrome(options=options)


def normalizar_texto(texto: str) -> str:
    texto = (texto or "").replace("\xa0", " ").lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def somente_digitos(valor: str) -> str:
    return re.sub(r"\D", "", valor or "")


def validar_numero_processo(numero: str) -> bool:
    return bool(re.fullmatch(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}", numero.strip()))


def aguardar_documento(driver: webdriver.Chrome) -> None:
    WebDriverWait(driver, TEMPO_ESPERA).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def texto_completo(driver: webdriver.Chrome) -> str:
    corpo = WebDriverWait(driver, TEMPO_ESPERA).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )
    visivel = corpo.text or ""
    dom = driver.execute_script("return document.body ? document.body.innerText : '';") or ""
    return f"{visivel}\n{dom}"


def contem_termo(texto: str, termos: Iterable[str]) -> bool:
    normalizado = normalizar_texto(texto)
    return any(normalizar_texto(t) in normalizado for t in termos)


def clicar(driver: webdriver.Chrome, elemento: WebElement) -> None:
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elemento)
    time.sleep(0.3)
    try:
        elemento.click()
    except (ElementClickInterceptedException, WebDriverException):
        driver.execute_script("arguments[0].click();", elemento)


def primeiro_visivel(driver: webdriver.Chrome, seletores: list[tuple[str, str]], timeout: int = 6) -> WebElement | None:
    for tipo, seletor in seletores:
        try:
            elemento = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((tipo, seletor))
            )
            if elemento.is_displayed():
                return elemento
        except (TimeoutException, NoSuchElementException, StaleElementReferenceException):
            continue
    return None


def salvar_diagnostico(driver: webdriver.Chrome, numero: str, texto: str = "") -> None:
    PASTA_DIAGNOSTICO.mkdir(parents=True, exist_ok=True)
    nome = numero.replace(".", "_").replace("-", "_")
    try:
        driver.save_screenshot(str(PASTA_DIAGNOSTICO / f"{nome}.png"))
    except WebDriverException:
        pass
    try:
        (PASTA_DIAGNOSTICO / f"{nome}.html").write_text(driver.page_source, encoding="utf-8")
        if texto:
            (PASTA_DIAGNOSTICO / f"{nome}.txt").write_text(texto, encoding="utf-8")
    except OSError:
        pass


# ============================================================
# PESQUISA E SELEÇÃO DO RESULTADO TST
# ============================================================

def localizar_campo_pesquisa(driver: webdriver.Chrome) -> WebElement:
    seletores = [
        (By.ID, "numeroProcesso"),
        (By.NAME, "numeroProcesso"),
        (By.CSS_SELECTOR, "input[placeholder*='processo' i]"),
        (By.CSS_SELECTOR, "input[aria-label*='processo' i]"),
        (By.CSS_SELECTOR, "input[type='text']"),
    ]
    campo = primeiro_visivel(driver, seletores, timeout=8)
    if campo is None:
        raise TimeoutException("Campo de número do processo não localizado.")
    return campo


def enviar_pesquisa(driver: webdriver.Chrome, campo: WebElement) -> None:
    botoes = [
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.XPATH, "//button[contains(translate(normalize-space(.),'PESQUISARCONSULTARBUSCAR','pesquisarconsultarbuscar'),'pesquisar')]"),
        (By.XPATH, "//button[contains(translate(normalize-space(.),'PESQUISARCONSULTARBUSCAR','pesquisarconsultarbuscar'),'consultar')]"),
        (By.XPATH, "//button[contains(translate(normalize-space(.),'PESQUISARCONSULTARBUSCAR','pesquisarconsultarbuscar'),'buscar')]"),
    ]
    botao = primeiro_visivel(driver, botoes, timeout=3)
    if botao is not None:
        clicar(driver, botao)
    else:
        campo.send_keys(Keys.ENTER)


def pesquisar_numero(driver: webdriver.Chrome, numero: str) -> str:
    driver.get(URL_CONSULTA)
    aguardar_documento(driver)
    campo = localizar_campo_pesquisa(driver)
    campo.click()
    campo.send_keys(Keys.CONTROL, "a")
    campo.send_keys(Keys.BACKSPACE)
    campo.send_keys(numero)

    html_antes = driver.page_source
    enviar_pesquisa(driver, campo)

    try:
        WebDriverWait(driver, TEMPO_ESPERA).until(
            lambda d: d.page_source != html_antes
            or contem_termo(texto_completo(d), TERMOS_NAO_ENCONTRADO)
        )
    except TimeoutException:
        pass
    time.sleep(2)
    return texto_completo(driver)


def obter_contexto_elemento(elemento: WebElement, niveis: int = 6) -> tuple[str, str]:
    """Retorna texto e HTML do elemento junto com seus ancestrais próximos."""
    textos: list[str] = []
    htmls: list[str] = []
    atual = elemento

    for _ in range(niveis + 1):
        try:
            texto = re.sub(r"\s+", " ", atual.text or "").strip()
            html = atual.get_attribute("outerHTML") or ""
            if texto:
                textos.append(texto)
            if html:
                htmls.append(html)
            atual = atual.find_element(By.XPATH, "..")
        except (NoSuchElementException, StaleElementReferenceException, WebDriverException):
            break

    return " ".join(textos), " ".join(htmls)


def score_candidato_tst(elemento: WebElement, numero: str) -> int:
    """
    Pontua o elemento usando também o card/linha pai.

    No portal, o botão pode conter apenas um ícone ou a palavra "Visualizar",
    enquanto o número e a indicação TST ficam no elemento ancestral.
    """
    try:
        texto_contexto, html_contexto = obter_contexto_elemento(elemento)
        href = elemento.get_attribute("href") or ""
        combinado = normalizar_texto(f"{texto_contexto} {html_contexto} {href}")
        numero_digitos = somente_digitos(numero)
        digitos_contexto = somente_digitos(combinado)

        score = 0
        if numero_digitos and numero_digitos in digitos_contexto:
            score += 10
        if "tribunal superior do trabalho" in combinado:
            score += 10
        if re.search(r"\btst\b", combinado):
            score += 8
        if any(x in combinado for x in ("3º grau", "3o grau", "3° grau", "terceiro grau", "3ª instância", "3a instancia")):
            score += 8
        if "detalhe-processo" in combinado:
            score += 4
        if re.search(r"/3(?:/|$|\?|#)", href):
            score += 5
        if elemento.is_displayed():
            score += 1

        return score
    except (StaleElementReferenceException, WebDriverException):
        return -1


def validar_pagina_detalhe_tst(driver: webdriver.Chrome, numero: str) -> bool:
    """
    Confirma um detalhe processual real.

    Como o domínio já é o PJe do TST, a página de detalhe nem sempre repete a
    sigla "TST" no texto. Por isso, validamos pelo número na URL/texto e por
    elementos típicos do detalhe, sem exigir que a palavra TST esteja visível.
    """
    texto = normalizar_texto(texto_completo(driver))
    url = normalizar_texto(driver.current_url)
    numero_digitos = somente_digitos(numero)

    numero_presente = (
        numero_digitos in somente_digitos(texto)
        or numero_digitos in somente_digitos(url)
    )

    indicadores_detalhe = [
        "movimentações processuais",
        "movimentacoes processuais",
        "dados do processo",
        "classe processual",
        "órgão julgador",
        "orgao julgador",
        "relator",
        "partes do processo",
        "detalhe-processo",
    ]
    detalhe_presente = any(i in texto or i in url for i in indicadores_detalhe)

    pagina_inicial = (
        url.rstrip("/") == normalizar_texto(URL_CONSULTA).rstrip("/")
        and not detalhe_presente
    )

    return numero_presente and detalhe_presente and not pagina_inicial


def trocar_para_nova_aba(driver: webdriver.Chrome, abas_antes: set[str]) -> None:
    novas = [h for h in driver.window_handles if h not in abas_antes]
    if novas:
        driver.switch_to.window(novas[-1])
        aguardar_documento(driver)
        time.sleep(1)


def abrir_resultado_tst(driver: webdriver.Chrome, numero: str) -> bool:
    """
    Localiza o resultado TST considerando o card inteiro, não apenas o texto
    do botão. Faz tentativas por links diretos, cards e elementos clicáveis.
    """
    numero_digitos = somente_digitos(numero)

    # Aguarda o número aparecer fora do campo de entrada ou surgir algum card.
    try:
        WebDriverWait(driver, TEMPO_ESPERA).until(
            lambda d: (
                len(d.find_elements(By.XPATH, "//a | //button | //*[@role='button'] | //*[@role='link']")) > 2
                and numero_digitos in somente_digitos(texto_completo(d))
            )
            or contem_termo(texto_completo(d), TERMOS_NAO_ENCONTRADO)
        )
    except TimeoutException:
        pass

    # 1) Links diretos de detalhe têm prioridade.
    links_detalhe = driver.find_elements(
        By.XPATH,
        "//a[contains(@href,'detalhe-processo')]"
    )
    candidatos: list[tuple[int, WebElement]] = []

    for elemento in links_detalhe:
        score = score_candidato_tst(elemento, numero)
        if score >= 11:
            candidatos.append((score + 5, elemento))

    # 2) Demais botões/links, usando o contexto dos ancestrais.
    elementos = driver.find_elements(
        By.XPATH,
        "//a | //button | //*[@role='button'] | //*[@role='link']"
    )
    for elemento in elementos:
        score = score_candidato_tst(elemento, numero)
        if score >= 11:
            candidatos.append((score, elemento))

    # Remove referências repetidas pelo HTML do elemento.
    unicos: list[tuple[int, WebElement]] = []
    vistos: set[str] = set()
    for score, elemento in sorted(candidatos, key=lambda x: x[0], reverse=True):
        try:
            chave = (elemento.get_attribute("outerHTML") or "")[:1000]
        except WebDriverException:
            continue
        if chave and chave not in vistos:
            vistos.add(chave)
            unicos.append((score, elemento))

    for _, elemento in unicos[:20]:
        try:
            url_antes = driver.current_url
            abas_antes = set(driver.window_handles)
            clicar(driver, elemento)
            time.sleep(1.5)
            trocar_para_nova_aba(driver, abas_antes)
            aguardar_documento(driver)

            if validar_pagina_detalhe_tst(driver, numero):
                return True

            # Se o elemento clicado era um botão interno, tenta links no card pai.
            try:
                card = elemento
                for _ in range(6):
                    links = card.find_elements(By.XPATH, ".//a[@href]")
                    for link in links:
                        href = link.get_attribute("href") or ""
                        if "detalhe-processo" in href and numero_digitos in somente_digitos(href + card.text):
                            abas_antes = set(driver.window_handles)
                            clicar(driver, link)
                            time.sleep(1.5)
                            trocar_para_nova_aba(driver, abas_antes)
                            aguardar_documento(driver)
                            if validar_pagina_detalhe_tst(driver, numero):
                                return True
                    card = card.find_element(By.XPATH, "..")
            except (NoSuchElementException, StaleElementReferenceException, WebDriverException):
                pass

            # Volta para a página de resultados apenas se realmente navegou.
            if driver.current_url != url_antes and len(driver.window_handles) == 1:
                driver.back()
                aguardar_documento(driver)
                time.sleep(1)
        except (StaleElementReferenceException, WebDriverException, TimeoutException):
            continue

    return False


# ============================================================
# MOVIMENTAÇÕES PROCESSUAIS
# ============================================================

def abrir_aba_movimentacoes(driver: webdriver.Chrome) -> bool:
    seletores = [
        (By.XPATH, "//*[@role='tab' and contains(translate(normalize-space(.),'ÇÕÁÀÂÃÉÊÍÓÔÚ','çõáàâãéêíóôú'),'movimentações processuais')]"),
        (By.XPATH, "//*[self::a or self::button][contains(translate(normalize-space(.),'ÇÕÁÀÂÃÉÊÍÓÔÚ','çõáàâãéêíóôú'),'movimentações processuais')]"),
        (By.XPATH, "//*[@role='tab' and contains(translate(normalize-space(.),'ÇÕÁÀÂÃÉÊÍÓÔÚ','çõáàâãéêíóôú'),'movimentações')]"),
        (By.XPATH, "//*[self::a or self::button][contains(translate(normalize-space(.),'ÇÕÁÀÂÃÉÊÍÓÔÚ','çõáàâãéêíóôú'),'movimentações')]"),
    ]
    aba = primeiro_visivel(driver, seletores, timeout=5)
    if aba is None:
        # Alguns layouts já exibem as movimentações sem aba separada.
        texto = normalizar_texto(texto_completo(driver))
        return "movimentações processuais" in texto or "movimentacoes processuais" in texto

    html_antes = driver.page_source
    clicar(driver, aba)
    try:
        WebDriverWait(driver, TEMPO_ESPERA).until(
            lambda d: d.page_source != html_antes
            or "movimenta" in normalizar_texto(texto_completo(d))
        )
    except TimeoutException:
        pass
    time.sleep(1.5)
    return True


def carregar_todas_movimentacoes(driver: webdriver.Chrome) -> None:
    expressoes = [
        "mostrar mais",
        "ver mais",
        "carregar mais",
        "mais movimentações",
        "mais movimentacoes",
        "exibir todas",
        "ver todas",
    ]

    for _ in range(20):
        clicou = False
        for expressao in expressoes:
            elementos = driver.find_elements(
                By.XPATH,
                "//*[self::button or self::a or @role='button']"
                f"[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÀÂÃÉÊÍÓÔÕÚÇ','abcdefghijklmnopqrstuvwxyzáàâãéêíóôõúç'),'{expressao}')]"
            )
            for elemento in elementos:
                try:
                    if elemento.is_displayed():
                        clicar(driver, elemento)
                        time.sleep(1)
                        clicou = True
                except (StaleElementReferenceException, WebDriverException):
                    continue

        altura_antes = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        altura_depois = driver.execute_script("return document.body.scrollHeight")
        if not clicou and altura_depois == altura_antes:
            break


def extrair_blocos_movimentacao(driver: webdriver.Chrome) -> list[str]:
    seletores = [
        "//*[contains(@class,'moviment')]",
        "//*[contains(@class,'timeline')]//*[(self::li or self::div)]",
        "//*[contains(@class,'andamento')]",
        "//*[@role='listitem']",
        "//table//tr",
    ]
    textos: list[str] = []
    vistos: set[str] = set()
    for xpath in seletores:
        for elemento in driver.find_elements(By.XPATH, xpath):
            try:
                texto = re.sub(r"\s+", " ", elemento.text or "").strip()
                if len(texto) < 8:
                    continue
                chave = normalizar_texto(texto)
                if chave not in vistos:
                    vistos.add(chave)
                    textos.append(texto)
            except (StaleElementReferenceException, WebDriverException):
                continue
    return textos


def analisar_movimentacoes(driver: webdriver.Chrome) -> tuple[str, str, str, str]:
    abrir_aba_movimentacoes(driver)
    carregar_todas_movimentacoes(driver)
    blocos = extrair_blocos_movimentacao(driver)
    texto_pagina = texto_completo(driver)
    universo = "\n".join(blocos) if blocos else texto_pagina
    universo_normalizado = normalizar_texto(universo)

    for termo in TERMOS_TRANSITO:
        termo_norm = normalizar_texto(termo)
        if termo_norm in universo_normalizado:
            trecho = next(
                (b for b in blocos if termo_norm in normalizar_texto(b)),
                extrair_trecho(universo, termo),
            )
            data = extrair_data(trecho)
            ultima = extrair_ultima_movimentacao(blocos, texto_pagina)
            return "Sim", termo, data, trecho or ultima

    return "Não localizado", "", "", extrair_ultima_movimentacao(blocos, texto_pagina)


def extrair_data(texto: str) -> str:
    match = re.search(r"\b\d{2}/\d{2}/\d{4}\b", texto or "")
    return match.group(0) if match else ""


def extrair_trecho(texto: str, termo: str, margem: int = 250) -> str:
    normal = normalizar_texto(texto)
    pos = normal.find(normalizar_texto(termo))
    if pos < 0:
        return ""
    limpo = re.sub(r"\s+", " ", texto).strip()
    return limpo[max(0, pos - margem): min(len(limpo), pos + len(termo) + margem)]


def extrair_ultima_movimentacao(blocos: list[str], texto_pagina: str) -> str:
    padrao_data = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
    candidatos = [b for b in blocos if padrao_data.search(b)]
    if candidatos:
        # Normalmente o PJe mostra a movimentação mais recente no topo.
        return candidatos[0][:1500]
    linhas = [re.sub(r"\s+", " ", l).strip() for l in texto_pagina.splitlines()]
    linhas = [l for l in linhas if padrao_data.search(l)]
    return linhas[0][:1500] if linhas else ""


def extrair_rotulo(texto: str, rotulos: list[str]) -> str:
    linhas = [re.sub(r"\s+", " ", l).strip() for l in (texto or "").splitlines() if l.strip()]
    for i, linha in enumerate(linhas):
        normal = normalizar_texto(linha)
        for rotulo in rotulos:
            r = normalizar_texto(rotulo)
            if normal.startswith(r):
                resto = re.sub(rf"^{re.escape(rotulo)}\s*:?-?\s*", "", linha, flags=re.I).strip()
                if resto and normalizar_texto(resto) != r:
                    return resto[:300]
                if i + 1 < len(linhas):
                    return linhas[i + 1][:300]
    return ""


# ============================================================
# RESULTADO
# ============================================================

def montar_resultado(
    processo: str,
    possui_tst: str,
    transito: str,
    situacao: str,
    fundamento: str = "",
    data_transito: str = "",
    classe: str = "",
    orgao: str = "",
    relator: str = "",
    ultima: str = "",
    trecho: str = "",
    url: str = "",
    erro: str = "",
) -> dict:
    return {
        "Processo": processo,
        "Possui processo no TST": possui_tst,
        "Trânsito em julgado no TST": transito,
        "Fundamento encontrado": fundamento,
        "Data do trânsito em julgado": data_transito,
        "Classe processual no TST": classe,
        "Órgão julgador": orgao,
        "Relator": relator,
        "Última movimentação identificada": ultima,
        "Trecho relevante": trecho,
        "Situação da consulta": situacao,
        "Data e hora da consulta": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "URL consultada": url,
        "Detalhes do erro": erro,
    }


def consultar_processo(driver: webdriver.Chrome, numero: str) -> dict:
    numero = numero.strip()
    if not validar_numero_processo(numero):
        return montar_resultado(numero, "Não consultado", "Não consultado", "Número em formato inválido")

    texto_pesquisa = pesquisar_numero(driver, numero)
    if contem_termo(texto_pesquisa, TERMOS_CAPTCHA):
        return montar_resultado(numero, "Não consultado", "Não consultado", "CAPTCHA identificado", url=driver.current_url)
    if contem_termo(texto_pesquisa, TERMOS_NAO_ENCONTRADO):
        return montar_resultado(numero, "Não", "Não se aplica", "Processo não encontrado no portal", url=driver.current_url)

    if not abrir_resultado_tst(driver, numero):
        salvar_diagnostico(driver, numero, texto_pesquisa)
        return montar_resultado(
            numero, "Não", "Não se aplica", "Nenhum resultado de terceira instância/TST localizado",
            url=driver.current_url,
        )

    if not validar_pagina_detalhe_tst(driver, numero):
        texto = texto_completo(driver)
        salvar_diagnostico(driver, numero, texto)
        return montar_resultado(
            numero, "Não confirmado", "Não analisado", "Página de detalhe do TST não confirmada",
            url=driver.current_url,
        )

    texto_detalhe = texto_completo(driver)
    transito, fundamento, data_transito, trecho = analisar_movimentacoes(driver)
    texto_final = texto_completo(driver)
    ultima = extrair_ultima_movimentacao(extrair_blocos_movimentacao(driver), texto_final)

    if SALVAR_DIAGNOSTICO_DE_SUCESSO:
        salvar_diagnostico(driver, numero, texto_final)

    return montar_resultado(
        processo=numero,
        possui_tst="Sim",
        transito=transito,
        fundamento=fundamento,
        data_transito=data_transito,
        classe=extrair_rotulo(texto_detalhe, ["Classe processual", "Classe"]),
        orgao=extrair_rotulo(texto_detalhe, ["Órgão julgador", "Orgao julgador"]),
        relator=extrair_rotulo(texto_detalhe, ["Relator", "Relatora"]),
        ultima=ultima,
        trecho=trecho,
        situacao="Consulta concluída — detalhe TST e movimentações analisados",
        url=driver.current_url,
    )


# ============================================================
# EXCEL — APENAS AO FINAL
# ============================================================

def salvar_excel_final(resultados: list[dict]) -> None:
    if not resultados:
        print("Nenhum resultado para salvar.")
        return

    colunas = [
        "Processo",
        "Possui processo no TST",
        "Trânsito em julgado no TST",
        "Fundamento encontrado",
        "Data do trânsito em julgado",
        "Classe processual no TST",
        "Órgão julgador",
        "Relator",
        "Última movimentação identificada",
        "Trecho relevante",
        "Situação da consulta",
        "Data e hora da consulta",
        "URL consultada",
        "Detalhes do erro",
    ]
    df = pd.DataFrame(resultados)
    for coluna in colunas:
        if coluna not in df.columns:
            df[coluna] = ""
    df = df[colunas]

    resumo = pd.DataFrame({
        "Indicador": [
            "Total de números analisados",
            "Processos encontrados no TST",
            "Processos sem resultado no TST",
            "Trânsito em julgado confirmado",
            "Trânsito não localizado",
            "Consultas com erro ou não confirmadas",
        ],
        "Quantidade": [
            len(df),
            int((df["Possui processo no TST"] == "Sim").sum()),
            int((df["Possui processo no TST"] == "Não").sum()),
            int((df["Trânsito em julgado no TST"] == "Sim").sum()),
            int((df["Trânsito em julgado no TST"] == "Não localizado").sum()),
            int(df["Situação da consulta"].str.contains("erro|captcha|não confirmada|não confirmado", case=False, na=False).sum()),
        ],
    })

    with pd.ExcelWriter(ARQUIVO_SAIDA, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resultados_TST", index=False)
        resumo.to_excel(writer, sheet_name="Resumo", index=False)

        cabecalho_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cabecalho_font = Font(color="FFFFFF", bold=True)

        ws = writer.book["Resultados_TST"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for celula in ws[1]:
            celula.fill = cabecalho_fill
            celula.font = cabecalho_font
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        larguras = {
            "A": 28, "B": 23, "C": 29, "D": 32, "E": 24, "F": 35, "G": 35,
            "H": 35, "I": 75, "J": 100, "K": 55, "L": 22, "M": 75, "N": 100,
        }
        for coluna, largura in larguras.items():
            ws.column_dimensions[coluna].width = largura
        for linha in ws.iter_rows(min_row=2):
            for celula in linha:
                celula.alignment = Alignment(vertical="top", wrap_text=True)

        ws_resumo = writer.book["Resumo"]
        ws_resumo.freeze_panes = "A2"
        ws_resumo.column_dimensions["A"].width = 45
        ws_resumo.column_dimensions["B"].width = 15
        for celula in ws_resumo[1]:
            celula.fill = cabecalho_fill
            celula.font = cabecalho_font
            celula.alignment = Alignment(horizontal="center")


# ============================================================
# EXECUÇÃO
# ============================================================

def main() -> None:
    resultados: list[dict] = []
    driver: webdriver.Chrome | None = None
    processos_unicos = list(dict.fromkeys(processos))

    print("=" * 78)
    print("CONSULTA TST — TERCEIRA INSTÂNCIA E MOVIMENTAÇÕES PROCESSUAIS")
    print(f"Processos informados: {len(processos)}")
    print(f"Processos únicos: {len(processos_unicos)}")
    print(f"Duplicados removidos: {len(processos) - len(processos_unicos)}")
    print("O Excel será criado somente no final.")
    print("=" * 78)

    try:
        driver = criar_driver()
        for indice, numero in enumerate(processos_unicos, start=1):
            print(f"\n[{indice}/{len(processos_unicos)}] {numero}")
            try:
                resultado = consultar_processo(driver, numero)
            except Exception as erro:
                if driver is not None:
                    salvar_diagnostico(driver, numero, texto_completo(driver) if driver.window_handles else "")
                resultado = montar_resultado(
                    numero,
                    "Não confirmado",
                    "Não analisado",
                    "Erro inesperado",
                    url=driver.current_url if driver else "",
                    erro=f"{type(erro).__name__}: {erro}\n{traceback.format_exc()}",
                )
            resultados.append(resultado)
            print(f"  TST: {resultado['Possui processo no TST']}")
            print(f"  Trânsito: {resultado['Trânsito em julgado no TST']}")
            print(f"  Situação: {resultado['Situação da consulta']}")
            if indice < len(processos_unicos):
                time.sleep(INTERVALO_ENTRE_PROCESSOS)
    finally:
        if driver is not None:
            driver.quit()

    # ÚNICO ponto de criação do Excel.
    salvar_excel_final(resultados)
    print("\n" + "=" * 78)
    print("PROCESSAMENTO FINALIZADO")
    print(f"Excel gerado: {ARQUIVO_SAIDA.resolve()}")
    print("=" * 78)


if __name__ == "__main__":
    main()
